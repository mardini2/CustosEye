# ruff: noqa: E501
"""
goal: flask web dashboard for CustosEye security monitoring system. provides a web interface
         for viewing live security events, process tree with trust scores, and file integrity
         monitoring. runs entirely locally without cloud dependencies.

what this app is responsible for:
- event ingestion: subscribes to the event bus from monitoring agents (process monitor, network
  scanner, integrity checker) and processes incoming events
- trust scoring: integrates with the CSC trust engine to evaluate process trustworthiness and
  attach verdicts (trusted, suspicious, malicious, etc.) to process events
- integrity monitoring: manages a watch list of files to monitor, computes baselines (SHA256 or
  mtime+size), detects changes, and provides diff views for changed files
- API endpoints: exposes REST endpoints for the frontend to fetch events, process tree, integrity
  targets, and file diffs
- event deduplication: uses fingerprint-based coalescing to prevent duplicate events from flooding
  the UI, with "worse event" promotion to ensure critical events aren't lost

how data flows through the app:
1. agents publish events to the event bus (process monitor, network scanner, integrity checker)
2. dashboard subscribes to the bus and drains events in a background thread
3. events are processed: rules engine tags them with levels/reasons, CSC engine scores processes,
   integrity events update target status
4. processed events are deduplicated and added to a bounded ring buffer (BUFFER)
5. frontend polls /api/events to fetch events from the buffer
6. frontend also polls /api/proctree to get the process tree built from process events
7. integrity operations (add target, rehash, view diff) go through dedicated API endpoints

major components and logic sections:
- event processing pipeline: drain_into_buffer() pulls events from bus, applies rules/trust scoring,
  deduplicates, and stores in BUFFER
- trust scoring integration: uses CSCTrustEngine to evaluate processes, with fast-paths for known
  system processes and optional name-based heuristics
- integrity target management: CRUD operations for watch list, automatic baseline computation,
  change detection, and status updates
- diff computation: chunk-based hashing for efficient change detection, text extraction from
  various file types (Office docs, PDFs, text files), character/word/line-level diff algorithms
- baseline snapshot storage: optional content-addressed storage for full file snapshots, with
  automatic pruning to stay within size limits
- Flask routes: main page, event API, process tree API, integrity management endpoints, diff
  viewer, file picker, static assets
"""

from __future__ import annotations

# --- standard library ---
import csv
import hashlib
import io
import json
import os
import re
import secrets
import sys
import threading
import time
import zipfile
from collections import deque
from collections.abc import Iterator
from pathlib import Path
from typing import Any, TypedDict, cast

# load environment variables from .env file before importing auth modules
try:
    from dotenv import load_dotenv

    load_dotenv()  # load .env file if it exists (required for auth secrets)
except ImportError:
    pass  # python-dotenv is optional, but required for .env support

# --- third-party ---
from flask import (
    Flask,
    jsonify,
    make_response,
    render_template,
    request,
    send_file,
)

# --- local/project imports (move these up here) ---
from agent.rules_engine import RulesEngine
from algorithm.csc_engine import CSCTrustEngine  # v2 under the hood
from dashboard.auth import SESSION_SECRET
from dashboard.auth_routes import register_auth_routes, require_auth
from dashboard.config import Config, load_config


# (now we can define classes/variables/etc.)
class RecentMeta(TypedDict, total=False):
    ref: dict[str, Any]
    seen: int
    last_seen: float


# single waitress optional block (we keep only this one, after all imports)
try:
    from waitress import serve as _serve  # type: ignore[import-untyped]

    HAVE_WAITRESS = True
except Exception:
    HAVE_WAITRESS = False
    _serve = None  # type: ignore

# load configuration and set up paths
CFG: Config = load_config()

ALLOW_MTIME_SNAPSHOTS: bool = bool(getattr(CFG, "allow_mtime_snapshots", True))

# ---------------- Config / paths ----------------
# extract all config paths for easy access throughout the app
BASE_DIR = CFG.base_dir

# session invalidation: generate a new session key on each program start
# this ensures all sessions from previous runs are invalidated when program restarts
# by changing the session secret key, all old session cookies become invalid
SESSION_KEY_FILE = BASE_DIR / "data" / ".session_key"
SESSION_START_TIME = time.time()

# generate new session key for this run (changes on each program start)
# this invalidates all sessions from previous program runs
_current_session_key = secrets.token_hex(32)
try:
    SESSION_KEY_FILE.parent.mkdir(parents=True, exist_ok=True)
    with open(SESSION_KEY_FILE, "w", encoding="utf-8") as f:
        f.write(_current_session_key)
except Exception:
    pass  # if we can't write, continue anyway (sessions will still work, just won't invalidate on restart)
RULES_PATH = str(CFG.rules_path)
CSC_WEIGHTS_PATH = str(CFG.csc_weights_path)
CSC_DB_PATH = str(CFG.csc_db_path)
INTEGRITY_TARGETS_PATH = str(CFG.integrity_targets_path)
SELF_SUPPRESS_PATH = str(CFG.self_suppress_path)


# load self-suppression list to filter out our own processes from the event stream
def _load_self_suppress() -> dict[str, set[str]]:
    try:
        with open(SELF_SUPPRESS_PATH, encoding="utf-8") as f:
            obj = json.load(f) or {}
    except Exception:
        obj = {}
    # normalize to lowercase sets
    return {
        "names": {s.lower() for s in obj.get("names", ["CustosEye.exe"])},
        "exes": {s.lower() for s in obj.get("exes", [])},
        "sha256": {s.lower() for s in obj.get("sha256", [])},
    }


SELF_SUPPRESS = _load_self_suppress()

# initialize rules engine and trust engine with config paths
_rules = RulesEngine(path=RULES_PATH)
_rules_mtime = os.path.getmtime(RULES_PATH) if os.path.exists(RULES_PATH) else 0.0
_csc = CSCTrustEngine(weights_path=CSC_WEIGHTS_PATH, db_path=CSC_DB_PATH)

# optional name-based trust heuristics (fast-path for known-good processes)
NAME_TRUST_PATH = str((BASE_DIR / "data" / "name_trust.json").resolve())
_name_trust: dict[str, tuple[str, str, float]] = {}
_name_trust_mtime: float = 0.0


# hot-reload name trust heuristics if the file changed
def _maybe_reload_name_trust() -> None:
    global _name_trust, _name_trust_mtime
    try:
        mtime = os.path.getmtime(NAME_TRUST_PATH)
    except OSError:
        mtime = 0.0
    if mtime > _name_trust_mtime:
        _name_trust_mtime = mtime
        try:
            with open(NAME_TRUST_PATH, encoding="utf-8") as f:
                obj = json.load(f)
            # normalize to {name: (verdict, cls, confidence)}
            nt: dict[str, tuple[str, str, float]] = {}
            for k, v in (obj or {}).items():
                if isinstance(v, list) and len(v) == 3:
                    verdict, cls, conf = v
                    nt[str(k).lower()] = (str(verdict), str(cls), float(conf))
            _name_trust = nt
        except Exception:
            _name_trust = {}


# fast-path: check if process name is in the trust map and apply verdict directly
def _maybe_promote_to_trusted(ev: dict) -> bool:
    nm = (ev.get("name") or "").lower()
    verdict = _name_trust.get(nm)
    if not verdict:
        return False
    v, cls, conf = verdict
    ev["csc"] = {
        "version": "v2",
        "verdict": v,
        "cls": cls,
        "confidence": conf,
        "reasons": ["name+parent heuristic"],
        "signals": {},
    }
    ev["csc_verdict"], ev["csc_class"], ev["csc_confidence"] = v, cls, conf
    ev["csc_reasons"] = ["name+parent heuristic"]
    return True


# hot-reload rules if the rules file changed (allows live rule updates)
def _maybe_reload_rules() -> None:
    global _rules, _rules_mtime
    try:
        mtime = os.path.getmtime(RULES_PATH)
    except OSError:
        mtime = 0.0
    if mtime > _rules_mtime:
        _rules_mtime = mtime
        _rules.rules = _rules._load_rules()


# ---------------- Buffers / indices ----------------
# bounded ring buffer for live events (oldest events drop when full)
BUFFER_MAX = CFG.buffer_max
BUFFER: deque[dict[str, Any]] = deque(maxlen=BUFFER_MAX)
# process index for building the process tree (PID -> process metadata)
PROC_INDEX: dict[int, dict[str, Any]] = {}


def _update_live_events_for_path(path: str, acceptance_text: str) -> None:
    """
    Update existing Live Events entries for a given path when baseline is accepted.
    Appends acceptance text to the existing reason, keeping the level as CRITICAL.
    Only updates integrity events related to file changes (not verification events).
    """
    try:
        path_lower = path.lower() if path else ""
        if not path_lower:
            return

        # update events in BUFFER
        for ev in BUFFER:
            ev_path = str(ev.get("path") or "").lower()
            ev_level = str(ev.get("level") or "").lower()
            ev_reason = str(ev.get("reason") or "")
            ev_reason_lower = ev_reason.lower()

            # match integrity events for this path that are critical and related to changes
            # skip if already accepted
            if (
                ev.get("source") == "integrity"
                and ev_path == path_lower
                and ev_level == "critical"
                and not ev.get("accepted", False)
                and (
                    "changed" in ev_reason_lower
                    or "mismatch" in ev_reason_lower
                    or "deleted" in ev_reason_lower
                    or "missing" in ev_reason_lower
                )
            ):
                # append acceptance text to existing reason, keep level as CRITICAL
                ev["reason"] = ev_reason + " — " + acceptance_text
                # preserve timestamp and level - don't change them
                # mark as accepted/approved
                ev["accepted"] = True

        # update events in RECENT_MAP
        for fp, rec in RECENT_MAP.items():
            ev_ref = rec.get("ref", {})
            ev_path = str(ev_ref.get("path") or "").lower()
            ev_level = str(ev_ref.get("level") or "").lower()
            ev_reason = str(ev_ref.get("reason") or "")
            ev_reason_lower = ev_reason.lower()

            # match integrity events for this path that are critical and related to changes
            # skip if already accepted
            if (
                ev_ref.get("source") == "integrity"
                and ev_path == path_lower
                and ev_level == "critical"
                and not ev_ref.get("accepted", False)
                and (
                    "changed" in ev_reason_lower
                    or "mismatch" in ev_reason_lower
                    or "deleted" in ev_reason_lower
                    or "missing" in ev_reason_lower
                )
            ):
                # append acceptance text to existing reason, keep level as CRITICAL
                ev_ref["reason"] = ev_reason + " — " + acceptance_text
                ev_ref["accepted"] = True
    except Exception:
        # do not break event processing on update errors
        pass


def _update_live_events_for_hash_verified(path: str) -> None:
    """
    update existing Live Events entries for a given path when hash is verified.
    replaces the reason with "✔ Hash verified", keeping the level as CRITICAL.
    only updates integrity events related to file changes (not verification events).
    """
    try:
        path_lower = path.lower() if path else ""
        if not path_lower:
            return

        # update events in BUFFER
        for ev in BUFFER:
            ev_path = str(ev.get("path") or "").lower()
            ev_level = str(ev.get("level") or "").lower()
            ev_reason = str(ev.get("reason") or "")
            ev_reason_lower = ev_reason.lower()

            # match integrity events for this path that are critical and related to changes
            # skip if already updated to "Hash verified"
            if (
                ev.get("source") == "integrity"
                and ev_path == path_lower
                and ev_level == "critical"
                and "hash verified" not in ev_reason_lower
                and (
                    "changed" in ev_reason_lower
                    or "mismatch" in ev_reason_lower
                    or "deleted" in ev_reason_lower
                    or "missing" in ev_reason_lower
                )
            ):
                # replace reason with "✔ Hash verified", keep level as CRITICAL
                ev["reason"] = "✔ Hash verified"
                # preserve timestamp and level - don't change them
                # mark as accepted/verified
                ev["accepted"] = True

        # update events in RECENT_MAP
        for fp, rec in RECENT_MAP.items():
            ev_ref = rec.get("ref", {})
            ev_path = str(ev_ref.get("path") or "").lower()
            ev_level = str(ev_ref.get("level") or "").lower()
            ev_reason = str(ev_ref.get("reason") or "")
            ev_reason_lower = ev_reason.lower()

            # match integrity events for this path that are critical and related to changes
            # skip if already updated to "Hash verified"
            if (
                ev_ref.get("source") == "integrity"
                and ev_path == path_lower
                and ev_level == "critical"
                and "hash verified" not in ev_reason_lower
                and (
                    "changed" in ev_reason_lower
                    or "mismatch" in ev_reason_lower
                    or "deleted" in ev_reason_lower
                    or "missing" in ev_reason_lower
                )
            ):
                # replace reason with "✔ Hash verified", keep level as CRITICAL
                ev_ref["reason"] = "✔ Hash verified"
                ev_ref["accepted"] = True
    except Exception:
        # do not break event processing on update errors
        pass


# --- dedupe/coalesce guard for live events ---
# deduplication system: prevents duplicate events from flooding the UI by coalescing similar
# events within a time window, but promotes "worse" events (higher severity, worse verdicts)
RECENT_TTL = 15.0  # seconds
RECENT_MAP: dict[str, RecentMeta] = {}  # fp -> {"ref": ev_dict, "seen": int, "last_seen": float}

# severity and verdict ranking for "worse event" promotion logic
_SEV_RANK = {"info": 0, "warning": 1, "critical": 2}
_VERDICT_RANK = {"trusted": 0, "caution": 1, "suspicious": 2, "malicious": 3, "unknown": 1}


# create a fingerprint from key event fields to identify duplicate events
def _fingerprint_base(ev: dict) -> str:
    key_fields = ("source", "reason", "pid", "name", "path", "rule")
    return json.dumps({k: ev.get(k) for k in key_fields}, sort_keys=True, ensure_ascii=False)


# check if the current event is "worse" than the previous one (higher severity, worse verdict)
# used to promote worse events even if they're duplicates
def _is_worse(prev: dict, cur: dict) -> bool:
    pl = _SEV_RANK.get(str(prev.get("level", "info")).lower(), 0)
    cl = _SEV_RANK.get(str(cur.get("level", "info")).lower(), 0)
    if cl > pl:
        return True

    pv = str(prev.get("csc_verdict", "unknown")).lower()
    cv = str(cur.get("csc_verdict", "unknown")).lower()
    if _VERDICT_RANK.get(cv, 1) > _VERDICT_RANK.get(pv, 1):
        return True
    if str(prev.get("csc_class", "")) != str(cur.get("csc_class", "")):
        return True
    try:
        if abs(float(cur.get("csc_confidence", 0)) - float(prev.get("csc_confidence", 0))) >= 0.05:
            return True
    except Exception:
        pass

    if (prev.get("source") == "integrity") or (cur.get("source") == "integrity"):
        if str(prev.get("last_result", "")) != str(cur.get("last_result", "")):
            return True
        if (
            str(cur.get("rule", "")).lower() == "sha256"
            and cur.get("sha256")
            and prev.get("sha256")
            and str(cur.get("sha256")).upper() != str(prev.get("sha256")).upper()
        ):
            return True

    return False


# deduplication logic: coalesce similar events or admit new/worse events to the buffer
def _coalesce_or_admit(ev: dict) -> bool:
    """
    returns True to append ev to BUFFER, False if merged into a recent one.
    keeps a single representative event within RECENT_TTL unless the new event is "worse".
    """
    now = time.time()
    # prune expired
    for k, rec in list(RECENT_MAP.items()):  # <— renamed from "entry"
        if now - float(rec.get("last_seen", 0)) > RECENT_TTL:
            RECENT_MAP.pop(k, None)

    fp = _fingerprint_base(ev)
    hit: RecentMeta | None = RECENT_MAP.get(fp)  # <— distinct name

    if hit is None:
        ev["seen"] = 1
        RECENT_MAP[fp] = {"ref": ev, "seen": 1, "last_seen": now}
        return True

    prev: dict[str, Any] = hit["ref"]

    # admit if the new one is worse (higher severity, worse verdict/class/conf)
    if _is_worse(prev, ev):
        ev["seen"] = 1
        RECENT_MAP[fp] = {"ref": ev, "seen": 1, "last_seen": now}
        return True

    # otherwise coalesce
    hit["seen"] = int(hit.get("seen", 1)) + 1
    hit["last_seen"] = now
    prev["seen"] = hit["seen"]
    prev["ts"] = ev.get("ts", now)  # keep the displayed timestamp fresh
    return False


# guardrails for event draining: limit how many events we process per API call and how long
# we spend draining to avoid blocking the web server
DRAIN_LIMIT_PER_CALL = CFG.drain_limit_per_call
DRAIN_DEADLINE_SEC = CFG.drain_deadline_sec

_DRAIN_LOCK = threading.Lock()  # to prevent concurrent drains


# ---------------- fan-out subscription plumbing ----------------
# subscribe to the event bus and return an iterator that yields events
def _bus_iterator(bus: Any) -> Iterator[dict[str, Any]]:
    """
    Return an iterator that yields events for this subscriber.
    Supports either:
      - bus.subscribe() -> iterator (preferred)
      - bus.iter_events() -> iterator (legacy)
    """
    if hasattr(bus, "subscribe"):
        it = bus.subscribe()
    else:
        it = bus.iter_events()
    # mypy: convince this is an Iterator[dict[str, Any]]
    return it  # type: ignore[return-value]


# integrity target management: read the watch list from JSON
def _read_integrity_targets() -> list[dict[str, Any]]:
    """
    back-compat:
      - if file is a JSON array: return it.
      - if file is an object with "targets": return that list.
      - if missing/unreadable: return [].
    also augments each row with "rule" (default sha256) and "last_result" if present.
    """
    try:
        with open(INTEGRITY_TARGETS_PATH, encoding="utf-8") as f:
            data = json.load(f)
        if isinstance(data, list):
            rows = cast(list[dict[str, Any]], data)
        elif isinstance(data, dict) and isinstance(data.get("targets"), list):
            rows = cast(list[dict[str, Any]], data["targets"])
        else:
            rows = []
    except Exception:
        rows = []
    # normalize fields for UI
    norm: list[dict[str, Any]] = []
    for r in rows:
        if not isinstance(r, dict):
            continue
        rr = dict(r)
        rr.setdefault("rule", "sha256" if rr.get("sha256") else "mtime+size")
        # last_result is not persisted by us; my agent may add it to events only.
        norm.append(rr)
    return norm


# update integrity target status when an integrity event comes in
def _update_integrity_target_from_event(ev: dict[str, Any]) -> None:
    """
    update integrity targets JSON when integrity events come in.
    note: Last Result is now only updated when user clicks Retest, not automatically from events.
    this gives users intentional control over status updates.
    """
    try:
        path = ev.get("path", "")
        if not path:
            return

        rows = _read_integrity_targets()
        target = next((r for r in rows if str(r.get("path") or "").lower() == path.lower()), None)
        if target is None:
            return  # not on watch list

        # only update hash if provided, but do not update last_result automatically
        # Last Result will be updated only when user clicks Retest
        if ev.get("actual"):
            target["sha256"] = ev.get("actual")
        elif ev.get("expected"):
            target["sha256"] = ev.get("expected")

        # only update status for OK/verified events (baseline accepted)
        reason = ev.get("reason", "").lower()

        # check for OK/verified status (baseline updated/accepted)
        if "baseline updated" in reason or "deletion accepted" in reason:
            if "baseline updated" in reason:
                target["last_result"] = "OK (baseline updated)"
            elif "deletion accepted" in reason:
                target["last_result"] = "OK (deletion accepted)"
            else:
                target["last_result"] = "OK (verified)"
            _write_integrity_targets(rows)
    except Exception:
        # do not break event processing on update errors
        pass


# write the integrity targets watch list back to JSON
def _write_integrity_targets(rows: list[dict[str, Any]]) -> None:
    """
    persist as a plain array to keep compatibility with your current file.
    """
    try:
        Path(INTEGRITY_TARGETS_PATH).parent.mkdir(parents=True, exist_ok=True)
        with open(INTEGRITY_TARGETS_PATH, "w", encoding="utf-8") as f:
            json.dump(rows, f, indent=2, ensure_ascii=False)
    except Exception:
        pass


# compute SHA256 hash of a file and return metadata (size, mtime, hash)
def _sha256_file(path: str) -> dict[str, Any]:
    # normalize (env vars, ~, quotes) but keep original text in "path"
    p = _norm_user_path(path)
    out: dict[str, Any] = {"path": path}
    try:
        st = os.stat(p)
        out["size"] = st.st_size
        out["mtime"] = int(st.st_mtime)
    except Exception as e:
        out["error"] = f"stat failed: {e}"
        return out
    try:
        h = hashlib.sha256()
        with open(p, "rb") as f:
            for chunk in iter(lambda: f.read(1024 * 1024), b""):
                h.update(chunk)
        out["sha256"] = h.hexdigest().upper()
    except Exception as e:
        out["error"] = f"hash failed: {e}"
    return out


# normalize file paths: expand environment variables, home directory, strip quotes
def _norm_user_path(p: str) -> str:
    # normalize, expand env vars (~, %WINDIR%), strip quotes, keep Unicode
    p = (p or "").strip().strip('"')
    p = os.path.expandvars(os.path.expanduser(p))
    return str(Path(p))


# ----- Integrity: chunk hashing + diff utilities -----
# chunk-based hashing: compute SHA256 for each fixed-size chunk to enable efficient
# region-level diff detection (only changed chunks need to be compared)
def _chunk_hashes(path: str, chunk_size: int = 4096) -> dict[str, Any]:
    # compute sha256 for each fixed-size chunk to enable region-level diffs
    p = _norm_user_path(path)  # normalize the path
    try:
        st = os.stat(p)  # stat to get size quickly
        total = int(st.st_size)  # total file size in bytes
    except Exception as e:
        return {"error": f"stat failed: {e}"}  # bubble error to caller

    hashes: list[str] = []  # per-chunk hex digests
    try:
        with open(p, "rb") as f:  # open file in binary
            while True:
                chunk = f.read(chunk_size)  # read a fixed-size block
                if not chunk:  # EOF
                    break
                h = hashlib.sha256()  # use sha256 for consistency
                h.update(chunk)  # hash the chunk
                hashes.append(h.hexdigest().upper())  # store uppercase hex
    except Exception as e:
        return {"error": f"read failed: {e}"}  # I/O errors return gracefully

    return {
        "algo": "sha256",  # hashing algorithm
        "chunk_size": int(chunk_size),  # size of each chunk in bytes
        "size": total,  # total size at time of baselining
        "hashes": hashes,  # list of per-chunk digests
    }


# utility: shorten long hex strings for display (show first N and last M chars)
def _short_hex(s: str, left: int = 8, right: int = 6) -> str:
    # shorten long hex strings for UI readability
    if not s:
        return ""
    if len(s) <= left + right + 1:
        return s
    return f"{s[:left]}…{s[-right:]}"


# merge consecutive chunk indexes into ranges for more compact diff display
def _merge_changed_chunks(changed_idxs: list[int]) -> list[tuple[int, int]]:
    # merge consecutive chunk indexes into ranges [start_idx, end_idx] inclusive
    if not changed_idxs:
        return []
    ranges: list[tuple[int, int]] = []
    start = prev = changed_idxs[0]
    for idx in changed_idxs[1:]:
        if idx == prev + 1:
            prev = idx
            continue
        ranges.append((start, prev))
        start = prev = idx
    ranges.append((start, prev))
    return ranges


# read a small preview of a changed region to show what the new content looks like
def _read_region_preview(path: str, start: int, length: int, cap: int = 64) -> bytes:
    # return up to `cap` bytes from the start of a region to preview "after" content
    p = _norm_user_path(path)
    try:
        with open(p, "rb") as f:
            f.seek(max(0, start))  # guard negative offsets
            return f.read(max(0, min(cap, length)))  # clamp to [0, cap]
    except Exception:
        return b""  # on error, no preview


# extract ZIP manifest for Office docs (docx/xlsx/pptx) to detect member-level changes
# this helps avoid false positives when Office docs are repacked but content is unchanged
def _zip_manifest(path: str) -> dict[str, dict[str, int]]:
    """
    If file is a ZIP, return { member_name: {size, crc, date} }.
    CRC is stored as unsigned int; date is an int like YYYYMMDD.
    If not ZIP or on error, return {}.
    """
    p = _norm_user_path(path)
    try:
        with open(p, "rb") as f:
            sig = f.read(4)
        if sig != b"PK\x03\x04":
            return {}
    except Exception:
        return {}

    out: dict[str, dict[str, int]] = {}
    try:
        with zipfile.ZipFile(p, "r") as zf:
            for zi in zf.infolist():
                # crc is already computed by the ZipInfo; convert to unsigned 32
                crc = zi.CRC & 0xFFFFFFFF
                # ZipInfo.date_time -> (Y, M, D, h, m, s)
                y, m, d = zi.date_time[:3]
                date = y * 10000 + m * 100 + d
                out[zi.filename] = {"size": zi.file_size, "crc": int(crc), "date": int(date)}
    except Exception:
        return {}
    return out


# detect file type from extension for human-readable diff summaries
def _nl_file_kind(path: str) -> str:
    p = (path or "").lower()
    if p.endswith((".pptx", ".ppt")):
        return "PowerPoint"
    if p.endswith((".docx", ".doc")):
        return "Word document"
    if p.endswith((".xlsx", ".xls")):
        return "Excel workbook"
    if p.endswith((".pdf",)):
        return "PDF"
    if p.endswith((".txt", ".log")):
        return "text file"
    if p.endswith(
        (
            ".py",
            ".js",
            ".ts",
            ".c",
            ".cpp",
            ".cs",
            ".java",
            ".rb",
            ".rs",
            ".go",
            ".php",
            ".sh",
            ".ps1",
            ".r",
            ".m",
            ".scala",
            ".kt",
        )
    ):
        return "source code"
    return "file"


# check if a file is likely a text file based on extension (for text extraction)
def _is_text_file(path: str) -> bool:
    """determine if a file is likely a text file based on extension."""
    p = (path or "").lower()
    text_extensions = (
        ".txt",
        ".log",
        ".md",
        ".json",
        ".xml",
        ".yaml",
        ".yml",
        ".ini",
        ".cfg",
        ".conf",
        ".py",
        ".js",
        ".ts",
        ".c",
        ".cpp",
        ".h",
        ".hpp",
        ".cs",
        ".java",
        ".rb",
        ".rs",
        ".go",
        ".php",
        ".sh",
        ".bash",
        ".ps1",
        ".bat",
        ".cmd",
        ".r",
        ".m",
        ".scala",
        ".kt",
        ".html",
        ".htm",
        ".css",
        ".scss",
        ".sql",
        ".csv",
        ".tsv",
        ".properties",
        ".env",
        ".gitignore",
        ".dockerfile",
        ".dockerignore",
        ".gitattributes",
        ".config",
        ".toml",
        ".lock",
        ".patch",
        ".diff",
        ".txt",
    )
    return p.endswith(text_extensions)


def _is_pdf_file(path: str) -> bool:
    """determine if a file is a PDF."""
    p = (path or "").lower()
    return p.endswith(".pdf")


def _is_config_file(path: str) -> bool:
    """determine if a file is a config file."""
    p = (path or "").lower()
    config_extensions = (
        ".json",
        ".xml",
        ".yaml",
        ".yml",
        ".ini",
        ".cfg",
        ".conf",
        ".toml",
        ".properties",
        ".env",
    )
    return p.endswith(config_extensions)


# read text file content with size limit to avoid loading huge files
def _read_text_file(path: str, max_size: int = 10 * 1024 * 1024) -> tuple[str | None, str]:
    """
    attempt to read a file as text.
    returns (content, error_message) where content is None on error.
    """
    try:
        p = _norm_user_path(path)
        st = os.stat(p)
        if st.st_size > max_size:
            return None, f"File too large ({st.st_size} bytes, max {max_size})"

        # try UTF-8 first
        with open(p, encoding="utf-8", errors="replace") as f:
            content = f.read()
        return content, ""
    except UnicodeDecodeError:
        # try other common encodings
        encodings = ["latin-1", "cp1252", "iso-8859-1"]
        for enc in encodings:
            try:
                with open(p, encoding=enc, errors="replace") as f:
                    content = f.read()
                return content, ""
            except Exception:
                continue
        return None, "Could not decode as text"
    except Exception as e:
        return None, f"Error reading file: {e}"


# extract readable text from PDF files for diff viewing (privacy-preserving)
def _extract_text_from_pdf(path: str) -> tuple[str | None, str]:
    """
    Extract readable text from PDF files.
    Returns (content, error_message) where content is None on error.
    Uses simple text extraction - for better results, consider using PyPDF2 or pdfplumber.
    """
    try:
        p = _norm_user_path(path)

        # simple PDF text extraction - look for readable text streams
        # this is a basic implementation; for production, consider using PyPDF2 or pdfplumber
        with open(p, "rb") as f:
            content = f.read()

        # try to extract text from PDF streams (basic approach)
        # look for text objects in PDF format
        import re

        # extract text from PDF streams (basic regex approach)
        # this is a simple fallback, proper PDF parsing would be better
        text_parts: list[str] = []

        # look for text in PDF streams
        stream_matches = re.findall(rb"stream\s+(.*?)\s+endstream", content, re.DOTALL)
        for stream in stream_matches[:20]:  # limit to first 20 streams
            try:
                # try to decode as text
                text = stream.decode("utf-8", errors="replace")
                # extract printable text
                text = re.sub(r"[^\x20-\x7E\n\r\t]", "", text)
                if len(text.strip()) > 10:  # only keep substantial text
                    text_parts.append(text.strip())
            except Exception:
                continue

        if text_parts:
            # join with line breaks
            return "\n".join(text_parts), ""

        return None, "No readable text found in PDF (consider using PyPDF2 for better extraction)"

    except Exception as e:
        return None, f"Error extracting text from PDF: {e}"


# extract embedded images from Office documents (docx/xlsx/pptx) for diff viewing
def _extract_images_from_office_doc(path: str) -> list[dict[str, Any]]:
    """
    extract image metadata from Office documents, including dimensions.
    returns list of image info dicts with name, path, size, hash, width, height.
    """
    images: list[dict[str, Any]] = []
    try:
        p = _norm_user_path(path)
        p_lower = (path or "").lower()

        if not p_lower.endswith((".docx", ".xlsx", ".pptx")):
            return images

        with zipfile.ZipFile(p, "r") as zf:
            # look for images in media folders
            if p_lower.endswith(".docx"):
                # Word: images in word/media/
                media_pattern = "word/media/"
                # Word stores image dimensions in document.xml relationships
                doc_xml = None
                try:
                    doc_xml = zf.read("word/document.xml").decode("utf-8", errors="replace")
                except Exception:
                    pass
            elif p_lower.endswith(".xlsx"):
                # Excel: images in xl/media/
                media_pattern = "xl/media/"
                doc_xml = None
            elif p_lower.endswith(".pptx"):
                # PowerPoint: images in ppt/media/
                media_pattern = "ppt/media/"
                doc_xml = None
            else:
                return images

            for name in zf.namelist():
                if name.startswith(media_pattern) and name.lower().endswith(
                    (".png", ".jpg", ".jpeg", ".gif", ".bmp", ".svg", ".wmf", ".emf")
                ):
                    try:
                        info = zf.getinfo(name)
                        # compute hash for the image
                        img_data = zf.read(name)
                        img_hash = hashlib.sha256(img_data).hexdigest()[
                            :16
                        ]  # short hash for display

                        # extract dimensions from document.xml for .docx
                        width = None
                        height = None

                        if p_lower.endswith(".docx") and doc_xml:
                            # find image relationships and extract dimensions
                            # Word stores dimensions in EMU (English Metric Units): 1 inch = 914400 EMU
                            # look for <wp:extent> or <a:ext> tags with cx/cy attributes
                            # search for relationships to this image
                            rel_pattern = r'<a:blip[^>]*r:embed="[^"]*"[^>]*>.*?<wp:extent[^>]*cx="(\d+)"[^>]*cy="(\d+)"[^>]*>'
                            matches = re.findall(rel_pattern, doc_xml, re.DOTALL)
                            if matches:
                                # use first match (most common case)
                                cx_emu, cy_emu = matches[0]
                                # convert EMU to pixels (assuming 96 DPI: 1 inch = 96 pixels = 914400 EMU)
                                width = round(int(cx_emu) / 914400 * 96)
                                height = round(int(cy_emu) / 914400 * 96)
                            else:
                                # try alternative pattern for inline images
                                alt_pattern = r'<a:blip[^>]*r:embed="[^"]*"[^>]*>.*?<a:ext[^>]*cx="(\d+)"[^>]*cy="(\d+)"[^>]*>'
                                alt_matches = re.findall(alt_pattern, doc_xml, re.DOTALL)
                                if alt_matches:
                                    cx_emu, cy_emu = alt_matches[0]
                                    width = round(int(cx_emu) / 914400 * 96)
                                    height = round(int(cy_emu) / 914400 * 96)

                        img_info = {
                            "name": os.path.basename(name),
                            "path": name,
                            "size": info.file_size,
                            "hash": img_hash,
                            "type": "image",
                        }

                        if width is not None and height is not None:
                            img_info["width"] = width
                            img_info["height"] = height

                        images.append(img_info)
                    except Exception:
                        continue

        return images
    except Exception:
        return images


# convert hex color code to friendly color name
def _hex_to_color_name(hex_code: str) -> str:
    """convert hex color code to friendly color name"""
    if not hex_code:
        return ""

    # normalize hex code (remove # if present, uppercase)
    hex_clean = hex_code.replace("#", "").upper()

    # common color mappings
    color_map = {
        "000000": "Black",
        "FFFFFF": "White",
        "FF0000": "Red",
        "00FF00": "Green",
        "0000FF": "Blue",
        "FFFF00": "Yellow",
        "FF00FF": "Magenta",
        "00FFFF": "Cyan",
        "800000": "Maroon",
        "008000": "Green",
        "000080": "Navy",
        "808000": "Olive",
        "800080": "Purple",
        "008080": "Teal",
        "C0C0C0": "Silver",
        "808080": "Gray",
        "FFA500": "Orange",
        "A02B93": "Purple",
        "E97132": "Orange",
        "156082": "Teal",
        "FFC0CB": "Pink",
        "A52A2A": "Brown",
        "DDA0DD": "Plum",
        "9370DB": "Medium Purple",
        "8B008B": "Dark Magenta",
        "4B0082": "Indigo",
    }

    # check exact match first
    if hex_clean in color_map:
        return color_map[hex_clean]

    # try to match common patterns by analyzing RGB components
    if len(hex_clean) == 6:
        try:
            r = int(hex_clean[0:2], 16)
            g = int(hex_clean[2:4], 16)
            b = int(hex_clean[4:6], 16)

            # Determine color based on dominant component
            max_val = max(r, g, b)
            min_val = min(r, g, b)
            diff = max_val - min_val

            # Gray scale detection
            if diff < 30:
                if max_val < 128:
                    return "Dark Gray"
                else:
                    return "Light Gray"

            # Teal/Cyan detection (blue-green) - check before generic blue/green
            # More lenient for colors like #156082 (r=21, g=96, b=130)
            if g > 80 and b > 100 and r < 100:
                return "Teal"

            # Color detection based on dominant component
            if r > g and r > b:
                if r > 200 and g < 100 and b < 100:
                    return "Red"
                elif r > 150 and g > 100 and b < 100:
                    return "Orange"
                elif r > 150 and g < 100 and b > 100:
                    return "Magenta"
                elif r > 100:
                    return "Pink"
            elif g > r and g > b:
                if g > 200 and r < 100 and b < 100:
                    return "Green"
                elif g > 150 and b > 100:
                    return "Cyan"
                elif g > 100:
                    return "Green"
            elif b > r and b > g:
                if b > 200 and r < 100 and g < 100:
                    return "Blue"
                elif b > 150 and r > 100:
                    return "Purple"
                elif b > 100:
                    return "Blue"

            # Brown detection (low brightness, red-orange)
            if max_val < 150 and r > g and r > b:
                return "Brown"
        except (ValueError, IndexError):
            pass

    # if no match, return empty string (will default to hex)
    return ""


# normalize text for comparison: Unicode punctuation, whitespace, apostrophes
def _normalize_text_for_diff(text: str) -> str:
    """
    normalize text for diff comparison to prevent spurious diffs.
    - converts curly apostrophes (U+2019) to straight apostrophes (')
    - normalizes non-breaking spaces to regular spaces
    - collapses repeated whitespace to single spaces
    returns normalized string for comparison only (original text is preserved in diffs).
    """
    import re

    if not text:
        return ""
    # convert curly apostrophes and quotes to straight ones
    normalized = text.replace("\u2019", "'")  # right single quotation mark → straight apostrophe
    normalized = normalized.replace(
        "\u2018", "'"
    )  # left single quotation mark → straight apostrophe
    normalized = normalized.replace("\u201C", '"')  # left double quotation mark → straight quote
    normalized = normalized.replace("\u201D", '"')  # right double quotation mark → straight quote
    # normalize non-breaking spaces to regular spaces
    normalized = normalized.replace("\u00A0", " ")  # non-breaking space → space
    normalized = normalized.replace("\u2009", " ")  # thin space → space
    normalized = normalized.replace("\u200A", " ")  # hair space → space
    # collapse repeated whitespace to single spaces (but preserve at least one space)
    normalized = re.sub(r"\s+", " ", normalized)
    return normalized.strip()


# calculate Jaccard similarity between two token sets
def _jaccard_similarity(tokens1: set[str], tokens2: set[str]) -> float:
    """
    calculate Jaccard similarity between two token sets.
    returns value between 0.0 (no overlap) and 1.0 (identical sets).
    used to detect when lines are similar enough to force word-level diff instead of full replacement.
    """
    if not tokens1 and not tokens2:
        return 1.0  # both empty = identical
    if not tokens1 or not tokens2:
        return 0.0  # one empty, one not = no similarity
    intersection = len(tokens1 & tokens2)
    union = len(tokens1 | tokens2)
    if union == 0:
        return 1.0
    return intersection / union


# tokenize text by Unicode word boundaries with grapheme clusters
def _tokenize_text(text: str) -> list[tuple[str, int, int]]:
    """
    Tokenize text by Unicode word boundaries, respecting grapheme clusters.
    Returns list of (token, start_pos, end_pos) tuples.
    Uses regex word boundaries which respect Unicode word boundaries (UAX#29).
    """
    import re

    if not text:
        return []
    tokens: list[tuple[str, int, int]] = []
    # Use \b for word boundaries - Python's re uses Unicode-aware word boundaries
    # This handles emoji, accents, zero-width joiners correctly
    # Match words (\b\w+\b) or non-whitespace sequences (\S+)
    for match in re.finditer(r"\b\w+\b|\S+", text):
        tokens.append((match.group(0), match.start(), match.end()))
    return tokens


# normalize style value for comparison (prevents phantom formatting changes)
def _normalize_style_value(attr_key: str, attr_val: Any) -> str:
    """
    normalize style value for comparison to prevent phantom formatting changes.
    converts colors to uppercase hex, normalizes boolean values, etc.
    returns normalized string value for comparison.
    """
    if attr_val is None:
        return ""

    val_str = str(attr_val).strip().lower()

    # normalize color values - convert to uppercase hex without # prefix
    if attr_key == "color":
        # remove # if present, convert to uppercase
        if val_str.startswith("#"):
            val_str = val_str[1:]
        # ensure it's a valid hex color (6 or 3 digits)
        if len(val_str) == 3:
            # expand 3-digit hex to 6-digit
            val_str = "".join(c * 2 for c in val_str)
        # pad to 6 digits if needed
        if len(val_str) < 6:
            val_str = val_str.ljust(6, "0")
        return val_str.upper()

    # normalize boolean attributes - convert to "true" or empty string
    if attr_key in ("bold", "italic", "strikethrough", "underline"):
        if val_str in ("true", "1", "yes", "on"):
            return "true"
        elif val_str in ("false", "0", "no", "off", ""):
            return ""
        # for strikethrough, handle "double" as a special case
        if attr_key == "strikethrough" and val_str == "double":
            return "double"
        return val_str

    # normalize other attributes - just lowercase and strip
    return val_str


# format attribute value for display (remove =true/false noise)
def _format_attr_for_display(attr_key: str, attr_val: Any) -> str:
    """Format attribute value for display, removing =true/false noise"""
    if attr_key == "color":
        # Color is already handled separately with hex and name
        return str(attr_val)
    elif attr_val in ("true", True, "1", 1):
        # Boolean true - just show the attribute name
        return attr_key
    elif attr_val in ("false", False, "0", 0):
        # Boolean false - shouldn't appear (means attribute removed)
        return ""
    else:
        # Other values - show as-is
        return f"{attr_key}: {attr_val}"


# extract text formatting (bold, italic, underline, strikethrough) from Office documents for diff viewing
def _extract_formatting_from_office_doc(path: str) -> dict[str, dict[str, Any]]:
    """
    extract formatting information (color, bold, italic, underline, strikethrough, etc.) from Office documents.
    returns a dict mapping text content to effective styles (after inheritance).
    this avoids false positives from run re-segmentation.
    """
    formatting_map: dict[str, dict[str, Any]] = {}
    try:
        p = _norm_user_path(path)
        p_lower = (path or "").lower()

        if not p_lower.endswith(".docx"):
            return formatting_map

        with zipfile.ZipFile(p, "r") as zf:
            try:
                doc_xml = zf.read("word/document.xml").decode("utf-8", errors="replace")

                # find all paragraphs first to get paragraph-level styles
                paragraph_pattern = r"<w:p[^>]*>(.*?)</w:p>"
                paragraphs = re.findall(paragraph_pattern, doc_xml, re.DOTALL)

                for para in paragraphs:
                    # get paragraph-level default styles (if any)
                    para_color = re.search(
                        r'<w:pPr[^>]*>.*?<w:color[^>]*w:val="([^"]+)"', para, re.DOTALL
                    )
                    # bold can be <w:b/> or <w:b w:val="true"/> or <w:b w:val="1"/>
                    para_bold = re.search(
                        r'<w:pPr[^>]*>.*?<w:b(?:\s+w:val="(?:true|1)")?[^>]*/>', para, re.DOTALL
                    )
                    para_italic = re.search(
                        r'<w:pPr[^>]*>.*?<w:i(?:\s+w:val="(?:true|1)")?[^>]*/>', para, re.DOTALL
                    )
                    para_strike = re.search(r"<w:pPr[^>]*>.*?<w:strike[^>]*/>", para, re.DOTALL)
                    para_dstrike = re.search(r"<w:pPr[^>]*>.*?<w:dstrike[^>]*/>", para, re.DOTALL)

                    # find all runs within this paragraph
                    run_pattern = r"<w:r[^>]*>(.*?)</w:r>"
                    runs = re.findall(run_pattern, para, re.DOTALL)

                    for run in runs:
                        # extract text - handle multi-codepoint graphemes (emojis) as single units
                        text_match = re.search(r"<w:t[^>]*>([^<]+)</w:t>", run)
                        if not text_match:
                            continue

                        text = text_match.group(1)
                        if not text.strip():
                            continue

                        # get run-level styles (override paragraph defaults)
                        # Word stores run properties in <w:rPr> tags - extract that first
                        rpr_match = re.search(r"<w:rPr[^>]*>(.*?)</w:rPr>", run, re.DOTALL)
                        rpr_content = rpr_match.group(1) if rpr_match else ""
                        # if no rPr tag, search in the entire run
                        search_in = rpr_content if rpr_content else run

                        # color: can be <w:color w:val="A02B93"/> (usually in rPr)
                        run_color = re.search(r'<w:color[^>]*w:val="([^"]+)"', search_in)

                        # bold can be <w:b/> or <w:b w:val="true"/> or <w:b w:val="1"/>
                        # but not <w:b w:val="false"/> or <w:b w:val="0"/>
                        run_bold_match = re.search(r"<w:b([^>]*)(?:/>|>)", search_in)
                        run_bold = False
                        if run_bold_match:
                            attrs = run_bold_match.group(1)
                            # check if val is explicitly false or 0
                            if not re.search(r'w:val="(?:false|0)"', attrs):
                                run_bold = True

                        # italic: can be <w:i/> or <w:i w:val="true"/> or <w:i w:val="1"/>
                        run_italic_match = re.search(r"<w:i([^>]*)(?:/>|>)", search_in)
                        run_italic = False
                        if run_italic_match:
                            attrs = run_italic_match.group(1)
                            # check if val is explicitly false or 0
                            if not re.search(r'w:val="(?:false|0)"', attrs):
                                run_italic = True

                        # underline: run-level only
                        run_underline = re.search(r'<w:u[^>]*w:val="([^"]+)"', search_in)

                        # strikethrough: can be <w:strike/> or <w:strike w:val="true"/>
                        # check for single strikethrough
                        run_strike_match = re.search(r"<w:strike([^>]*)(?:/>|>)", search_in)
                        run_strike = False
                        if run_strike_match:
                            attrs = run_strike_match.group(1)
                            # check if val is explicitly false or 0
                            if not re.search(r'w:val="(?:false|0)"', attrs):
                                run_strike = True

                        # double strikethrough: can be <w:dstrike/> or <w:dstrike w:val="true"/>
                        run_dstrike_match = re.search(r"<w:dstrike([^>]*)(?:/>|>)", search_in)
                        run_dstrike = False
                        if run_dstrike_match:
                            attrs = run_dstrike_match.group(1)
                            # check if val is explicitly false or 0
                            if not re.search(r'w:val="(?:false|0)"', attrs):
                                run_dstrike = True

                        # determine effective styles (run overrides paragraph)
                        effective_styles: dict[str, str] = {}

                        # color: run overrides paragraph
                        if run_color:
                            effective_styles["color"] = run_color.group(1)
                        elif para_color:
                            effective_styles["color"] = para_color.group(1)

                        # bold: run overrides paragraph
                        if run_bold:
                            effective_styles["bold"] = "true"
                        elif para_bold:
                            effective_styles["bold"] = "true"

                        # italic: run overrides paragraph
                        if run_italic:
                            effective_styles["italic"] = "true"
                        elif para_italic:
                            effective_styles["italic"] = "true"

                        # underline: run-level only
                        if run_underline:
                            effective_styles["underline"] = run_underline.group(1)

                        # strikethrough: run overrides paragraph
                        # single strikethrough (<w:strike/>)
                        if run_strike:
                            effective_styles["strikethrough"] = "true"
                        elif para_strike:
                            effective_styles["strikethrough"] = "true"

                        # double strikethrough (<w:dstrike/>)
                        if run_dstrike:
                            effective_styles["strikethrough"] = "double"
                        elif para_dstrike:
                            effective_styles["strikethrough"] = "double"

                        # store by text content
                        # IMPORTANT: store ALL formatting for each text occurrence
                        # don't merge - let the comparison logic handle merging when matching to unchanged text
                        text_key = text

                        # store formatting - handle multiple occurrences of same text
                        # IMPORTANT: preserve ALL formatting attributes from ALL occurrences
                        # This ensures color, strikethrough, italic are not lost
                        if text_key in formatting_map:
                            existing = formatting_map[text_key]
                            existing_styles = existing.get("styles", {})
                            # merge styles: combine all attributes from both occurrences
                            # use union of attributes - if attribute exists in either, include it
                            merged_styles = dict(existing_styles)

                            # add all attributes from new occurrence
                            for key, val in effective_styles.items():
                                # if attribute already exists, keep both values (comparison will handle it)
                                # for boolean attributes (bold, italic, strikethrough), if either has it, keep it
                                if key in merged_styles:
                                    # attribute exists in both - if values differ, keep the new one
                                    # (this handles cases where formatting changed)
                                    if key in ("bold", "italic", "strikethrough", "underline"):
                                        # for boolean attributes, if new occurrence has it, use it
                                        if val == "true" or val == "double":
                                            merged_styles[key] = val
                                    else:
                                        # for other attributes (like color), prefer the new value
                                        merged_styles[key] = val
                                else:
                                    # new attribute - add it
                                    merged_styles[key] = val

                            # also preserve attributes from existing that aren't in new
                            # (this ensures we don't lose formatting when text appears with different formatting)
                            for key, val in existing_styles.items():
                                if key not in merged_styles:
                                    merged_styles[key] = val
                                # for boolean attributes, if existing has it but new doesn't, check if we should keep it
                                elif key in ("bold", "italic", "strikethrough", "underline"):
                                    # if existing had the attribute and new doesn't explicitly set it,
                                    # we might want to keep it, but actually the new occurrence should override
                                    # so we'll use the merged value from above
                                    pass

                            formatting_map[text_key] = {
                                "text": text_key,
                                "styles": merged_styles,
                                "type": "formatting",
                            }
                        else:
                            # first occurrence - store as-is
                            formatting_map[text_key] = {
                                "text": text_key,
                                "styles": effective_styles,
                                "type": "formatting",
                            }
            except Exception:
                pass

        return formatting_map
    except Exception:
        return formatting_map


# extract readable text from Office documents for diff viewing
def _extract_text_from_office_doc(path: str) -> tuple[str | None, str]:
    """
    extract readable text from Office documents (.docx, .xlsx, .pptx).
    returns (content, error_message) where content is None on error.
    Office documents are ZIP files containing XML.
    also preserves whitespace (Tab/Enter) for proper diffing.
    """
    try:
        p = _norm_user_path(path)
        p_lower = (path or "").lower()

        if not p_lower.endswith((".docx", ".xlsx", ".pptx")):
            return None, "Not an Office document"

        text_parts: list[str] = []

        with zipfile.ZipFile(p, "r") as zf:
            if p_lower.endswith(".docx"):
                # Word: text is in word/document.xml
                try:
                    doc_xml = zf.read("word/document.xml").decode("utf-8", errors="replace")

                    # better approach: extract paragraphs and their text runs
                    # find all paragraphs first
                    paragraph_pattern = r"<w:p[^>]*>(.*?)</w:p>"
                    paragraphs = re.findall(paragraph_pattern, doc_xml, re.DOTALL)

                    if paragraphs:
                        para_texts = []
                        for para in paragraphs:
                            # extract all text runs within this paragraph
                            # preserve whitespace (Tab, Enter) by checking for <w:br/> and <w:tab/>
                            text_runs = re.findall(r"<w:t[^>]*>([^<]+)</w:t>", para)
                            # check for line breaks and tabs
                            has_break = re.search(r"<w:br[^>]*/>", para)
                            has_tab = re.search(r"<w:tab[^>]*/>", para)

                            if text_runs:
                                # join text runs within a paragraph (preserves formatting within para)
                                para_text = "".join(text_runs)
                                # decode XML entities
                                para_text = (
                                    para_text.replace("&lt;", "<")
                                    .replace("&gt;", ">")
                                    .replace("&amp;", "&")
                                )
                                # preserve tabs and line breaks
                                if has_tab:
                                    para_text = "\t" + para_text  # add tab at start if present
                                if has_break:
                                    para_text = para_text + "\n"  # add line break if present
                                if para_text.strip() or has_tab or has_break:
                                    para_texts.append(para_text)

                        if para_texts:
                            # join paragraphs with newlines to preserve structure
                            # do not strip to preserve whitespace
                            text = "\n".join(para_texts)
                            text_parts.append(f"[Word Document Text]\n{text}\n")
                    else:
                        # fallback: extract all text from <w:t> tags
                        text_matches = re.findall(r"<w:t[^>]*>([^<]+)</w:t>", doc_xml)
                        if text_matches:
                            # decode XML entities
                            text = "".join(text_matches)
                            text = (
                                text.replace("&lt;", "<").replace("&gt;", ">").replace("&amp;", "&")
                            )
                            # add newlines at paragraph boundaries (approximate)
                            text = re.sub(r"</w:p>", "\n", text)
                            text = re.sub(r"\n\s*\n", "\n", text)  # clean up multiple newlines
                            if text.strip():
                                text_parts.append(f"[Word Document Text]\n{text.strip()}\n")
                except Exception as e:
                    text_parts.append(f"[Word Document - Error extracting text: {e}]\n")

            elif p_lower.endswith(".xlsx"):
                # Excel: text is in xl/sharedStrings.xml and sheet data
                try:
                    # get shared strings
                    try:
                        shared_strings_xml = zf.read("xl/sharedStrings.xml").decode(
                            "utf-8", errors="replace"
                        )
                        # extract text from <t> tags (Excel text elements)
                        text_matches = re.findall(r"<t[^>]*>([^<]+)</t>", shared_strings_xml)
                        if text_matches:
                            text_parts.append(
                                "[Excel Shared Strings]\n" + "\n".join(text_matches) + "\n"
                            )
                    except Exception:
                        pass

                    # get sheet data (first sheet)
                    try:
                        sheet_list = [
                            name
                            for name in zf.namelist()
                            if name.startswith("xl/worksheets/sheet") and name.endswith(".xml")
                        ]
                        if sheet_list:
                            # read first sheet
                            sheet_xml = zf.read(sheet_list[0]).decode("utf-8", errors="replace")
                            # extract cell values with references
                            cell_matches = re.findall(
                                r'<c r="([^"]+)"[^>]*><v>([^<]+)</v></c>', sheet_xml
                            )
                            if cell_matches:
                                cell_data = [
                                    f"{ref}: {val}" for ref, val in cell_matches[:50]
                                ]  # Limit to 50 cells
                                text_parts.append(
                                    "[Excel Sheet Data (first 50 cells)]\n"
                                    + "\n".join(cell_data)
                                    + "\n"
                                )
                    except Exception:
                        pass
                except Exception as e:
                    text_parts.append(f"[Excel Document - Error extracting text: {e}]\n")

            elif p_lower.endswith(".pptx"):
                # PowerPoint: text is in ppt/slides/*.xml
                try:
                    slide_files = [
                        name
                        for name in zf.namelist()
                        if name.startswith("ppt/slides/slide") and name.endswith(".xml")
                    ]
                    slide_files.sort()

                    for slide_file in slide_files[:10]:  # limit to first 10 slides
                        try:
                            slide_xml = zf.read(slide_file).decode("utf-8", errors="replace")
                            # extract text from <a:t> tags (text in PowerPoint)
                            text_matches = re.findall(r"<a:t[^>]*>([^<]+)</a:t>", slide_xml)
                            if text_matches:
                                slide_num = slide_file.split("/")[-1].replace(".xml", "")
                                text_parts.append(
                                    f"[Slide {slide_num}]\n" + "\n".join(text_matches) + "\n"
                                )
                        except Exception:
                            continue
                except Exception as e:
                    text_parts.append(f"[PowerPoint Document - Error extracting text: {e}]\n")

        if text_parts:
            return "\n".join(text_parts), ""
        return None, "No text content found in Office document"

    except zipfile.BadZipFile:
        return None, "Not a valid ZIP file (Office document)"
    except Exception as e:
        return None, f"Error extracting text from Office document: {e}"


# split text into words for word-level diffing (better readability than character-level)
def _split_into_words(text: str) -> list[tuple[str, int, int]]:
    """
    split text into words, preserving whitespace.
    returns list of (word, start_pos, end_pos) tuples.
    """
    import re

    words: list[tuple[str, int, int]] = []
    for match in re.finditer(r"\S+|\s+", text):
        words.append((match.group(0), match.start(), match.end()))
    return words


# compute character-level or word-level diff for a single line (used in line diffs)
def _compute_char_diff(
    old_line: str, new_line: str, word_level: bool = True
) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    """
    compute word-level or character-level diff for a single line pair.
    returns (old_parts, new_parts) where each part is {'type': 'equal'|'removed'|'added', 'text': str}

    when word_level=True, groups changes by words for better readability.
    uses normalization to handle Unicode punctuation variants (curly vs straight apostrophes).
    """
    from difflib import SequenceMatcher

    # initialize result lists
    old_parts: list[dict[str, Any]] = []
    new_parts: list[dict[str, Any]] = []

    if word_level:
        # word-level diffing with better matching: normalize words for comparison but preserve original text
        old_words = _split_into_words(old_line)
        new_words = _split_into_words(new_line)

        # create normalized word lists for matching: normalize Unicode punctuation and strip punctuation for comparison
        # this helps match words like "what's" and "whats" or "threat" and "threat," as the same word
        import re

        def normalize_word_for_matching(word):
            # first normalize Unicode punctuation (curly apostrophes → straight)
            normalized = word.replace("\u2019", "'").replace("\u2018", "'")
            # strip all punctuation and whitespace for comparison, convert to lowercase
            # this allows "threat" and "threat," to be matched as the same word
            # also allows "what's" and "whats" to be matched
            word_clean = re.sub(r"[^\w]", "", normalized.lower())
            return word_clean if word_clean else normalized.lower()

        old_word_texts = [w[0] for w in old_words]
        new_word_texts = [w[0] for w in new_words]
        old_normalized = [normalize_word_for_matching(w) for w in old_word_texts]
        new_normalized = [normalize_word_for_matching(w) for w in new_word_texts]

        # use autojunk=False for better matching of similar sequences
        # SequenceMatcher will match words based on normalized comparison
        # this ensures apostrophe variants (what's vs whats) are matched correctly
        matcher = SequenceMatcher(None, old_normalized, new_normalized, autojunk=False)

        for tag, i1, i2, j1, j2 in matcher.get_opcodes():
            if tag == "equal":
                # reconstruct text from original line using word positions to preserve spacing
                if i1 < len(old_words) and i2 > 0:
                    start_pos = old_words[i1][1]
                    end_pos = old_words[i2 - 1][2]
                    old_text = old_line[start_pos:end_pos]
                else:
                    old_text = "".join(old_word_texts[i1:i2])

                if j1 < len(new_words) and j2 > 0:
                    start_pos = new_words[j1][1]
                    end_pos = new_words[j2 - 1][2]
                    new_text = new_line[start_pos:end_pos]
                else:
                    new_text = "".join(new_word_texts[j1:j2])

                old_parts.append({"type": "equal", "text": old_text})
                new_parts.append({"type": "equal", "text": new_text})
            elif tag == "delete":
                # reconstruct deleted text from original line
                if i1 < len(old_words) and i2 > 0:
                    start_pos = old_words[i1][1]
                    end_pos = old_words[i2 - 1][2]
                    text = old_line[start_pos:end_pos]
                else:
                    text = "".join(old_word_texts[i1:i2])
                old_parts.append({"type": "removed", "text": text})
            elif tag == "insert":
                # reconstruct inserted text from new line
                if j1 < len(new_words) and j2 > 0:
                    start_pos = new_words[j1][1]
                    end_pos = new_words[j2 - 1][2]
                    text = new_line[start_pos:end_pos]
                else:
                    text = "".join(new_word_texts[j1:j2])
                new_parts.append({"type": "added", "text": text})
            elif tag == "replace":
                # reconstruct replaced text from both lines
                if i1 < len(old_words) and i2 > 0:
                    start_pos = old_words[i1][1]
                    end_pos = old_words[i2 - 1][2]
                    old_text = old_line[start_pos:end_pos]
                else:
                    old_text = "".join(old_word_texts[i1:i2])

                if j1 < len(new_words) and j2 > 0:
                    start_pos = new_words[j1][1]
                    end_pos = new_words[j2 - 1][2]
                    new_text = new_line[start_pos:end_pos]
                else:
                    new_text = "".join(new_word_texts[j1:j2])

                old_parts.append({"type": "removed", "text": old_text})
                new_parts.append({"type": "added", "text": new_text})
    else:
        # character-level diffing (fallback)
        matcher = SequenceMatcher(None, old_line, new_line)

        for tag, i1, i2, j1, j2 in matcher.get_opcodes():
            if tag == "equal":
                old_parts.append({"type": "equal", "text": old_line[i1:i2]})
                new_parts.append({"type": "equal", "text": new_line[j1:j2]})
            elif tag == "delete":
                old_parts.append({"type": "removed", "text": old_line[i1:i2]})
            elif tag == "insert":
                new_parts.append({"type": "added", "text": new_line[j1:j2]})
            elif tag == "replace":
                old_parts.append({"type": "removed", "text": old_line[i1:i2]})
                new_parts.append({"type": "added", "text": new_line[j1:j2]})

    return old_parts, new_parts


# compute line-by-line diff with character-level highlighting for modified lines
# two-stage approach: Stage A (line anchors) + Stage B (word-level within changed lines)
def _compute_line_diff(old_lines: list[str], new_lines: list[str]) -> list[dict[str, Any]]:
    """
    compute a line-by-line diff with character-level highlighting for modified lines.
    uses two-stage approach:
    - Stage A: line-level anchors using LCS to align unchanged lines/blocks
    - Stage B: word-level diff within changed lines, with Jaccard similarity check
      to prevent full-line replacements when only a few words differ

    returns a list of diff segments with type: 'equal', 'added', 'removed', 'modified'
    for modified lines, includes character-level differences.
    """
    from difflib import SequenceMatcher

    # Stage A: line-level anchors using LCS (SequenceMatcher)
    # normalize lines for comparison to handle Unicode punctuation variants
    old_normalized = [_normalize_text_for_diff(line) for line in old_lines]
    new_normalized = [_normalize_text_for_diff(line) for line in new_lines]

    matcher = SequenceMatcher(None, old_normalized, new_normalized, autojunk=False)
    diff_segments: list[dict[str, Any]] = []

    for tag, i1, i2, j1, j2 in matcher.get_opcodes():
        if tag == "equal":
            # include equal lines (unchanged blocks - Stage A anchors)
            diff_segments.append(
                {
                    "type": "equal",
                    "old_start": i1 + 1,  # 1-indexed for display
                    "old_end": i2,
                    "new_start": j1 + 1,
                    "new_end": j2,
                    "old_lines": old_lines[i1:i2],
                    "new_lines": new_lines[j1:j2],
                }
            )
        elif tag == "delete":
            # lines removed
            diff_segments.append(
                {
                    "type": "removed",
                    "old_start": i1 + 1,
                    "old_end": i2,
                    "old_lines": old_lines[i1:i2],
                    "new_lines": [],
                }
            )
        elif tag == "insert":
            # lines added
            diff_segments.append(
                {
                    "type": "added",
                    "new_start": j1 + 1,
                    "new_end": j2,
                    "old_lines": [],
                    "new_lines": new_lines[j1:j2],
                }
            )
        elif tag == "replace":
            # Stage B: within changed lines, use word-level diff with Jaccard similarity check
            # this prevents full-line replacements when only a few words differ
            old_section = old_lines[i1:i2]
            new_section = new_lines[j1:j2]

            # if same number of lines, check each line individually
            if len(old_section) == len(new_section):
                # process each line pair individually to detect which lines actually changed
                for line_idx, (old_line, new_line) in enumerate(zip(old_section, new_section)):
                    # check if lines are actually identical (after normalization)
                    old_norm = _normalize_text_for_diff(old_line)
                    new_norm = _normalize_text_for_diff(new_line)
                    if old_norm == new_norm:
                        # lines are identical after normalization, mark as equal (no stripe)
                        diff_segments.append(
                            {
                                "type": "equal",
                                "old_start": i1 + line_idx + 1,
                                "old_end": i1 + line_idx + 1,
                                "new_start": j1 + line_idx + 1,
                                "new_end": j1 + line_idx + 1,
                                "old_lines": [old_line],
                                "new_lines": [new_line],
                            }
                        )
                    else:
                        # lines are different - ALWAYS use word-level diff to show only changed words
                        # this ensures that adding/removing a single word only highlights that word, not the whole line
                        # compute word-level diff for all line changes (not just similar ones)
                        old_parts, new_parts = _compute_char_diff(
                            old_line, new_line, word_level=True
                        )

                        # determine change type for stripe color: check if it's pure addition, pure removal, or mixed
                        has_removed = any(part.get("type") == "removed" for part in old_parts)
                        has_added = any(part.get("type") == "added" for part in new_parts)

                        # determine change_type for stripe: "addition_only", "removal_only", or "mixed"
                        # segment type stays as "modified" since line exists in both versions
                        if has_added and not has_removed:
                            change_type = "addition_only"
                        elif has_removed and not has_added:
                            change_type = "removal_only"
                        else:
                            change_type = "mixed"  # both additions and removals

                        diff_segments.append(
                            {
                                "type": "modified",  # line exists in both versions - use word-level diff
                                "change_type": change_type,  # used for stripe color
                                "old_start": i1 + line_idx + 1,
                                "old_end": i1 + line_idx + 1,
                                "new_start": j1 + line_idx + 1,
                                "new_end": j1 + line_idx + 1,
                                "old_lines": [old_line],
                                "new_lines": [new_line],
                                "old_char_diffs": [old_parts],  # word-level diff parts
                                "new_char_diffs": [new_parts],  # word-level diff parts
                            }
                        )
            else:
                # different number of lines - handle insertions/deletions
                # use a more sophisticated alignment to handle cases like:
                # baseline: ["add", QUOTE_LINE]
                # current: ["add", "hello", QUOTE_LINE]
                # we want: "add" = equal, "hello" = added, QUOTE_LINE = equal

                # try to align lines using normalized comparison
                # this helps when a line is inserted in the middle
                old_normalized_section = [_normalize_text_for_diff(line) for line in old_section]
                new_normalized_section = [_normalize_text_for_diff(line) for line in new_section]

                # use SequenceMatcher to align lines even when counts differ
                section_matcher = SequenceMatcher(
                    None, old_normalized_section, new_normalized_section, autojunk=False
                )

                for sect_tag, sect_i1, sect_i2, sect_j1, sect_j2 in section_matcher.get_opcodes():
                    if sect_tag == "equal":
                        # lines match - mark as equal
                        for line_idx in range(sect_i2 - sect_i1):
                            old_line = old_section[sect_i1 + line_idx]
                            new_line = new_section[sect_j1 + line_idx]
                            diff_segments.append(
                                {
                                    "type": "equal",
                                    "old_start": i1 + sect_i1 + line_idx + 1,
                                    "old_end": i1 + sect_i1 + line_idx + 1,
                                    "new_start": j1 + sect_j1 + line_idx + 1,
                                    "new_end": j1 + sect_j1 + line_idx + 1,
                                    "old_lines": [old_line],
                                    "new_lines": [new_line],
                                }
                            )
                    elif sect_tag == "delete":
                        # lines removed
                        for line_idx in range(sect_i2 - sect_i1):
                            old_line = old_section[sect_i1 + line_idx]
                            diff_segments.append(
                                {
                                    "type": "removed",
                                    "old_start": i1 + sect_i1 + line_idx + 1,
                                    "old_end": i1 + sect_i1 + line_idx + 1,
                                    "old_lines": [old_line],
                                    "new_lines": [],
                                }
                            )
                    elif sect_tag == "insert":
                        # lines added
                        for line_idx in range(sect_j2 - sect_j1):
                            new_line = new_section[sect_j1 + line_idx]
                            diff_segments.append(
                                {
                                    "type": "added",
                                    "new_start": j1 + sect_j1 + line_idx + 1,
                                    "new_end": j1 + sect_j1 + line_idx + 1,
                                    "old_lines": [],
                                    "new_lines": [new_line],
                                }
                            )
                    elif sect_tag == "replace":
                        # lines replaced - apply Jaccard similarity check
                        for line_idx in range(min(sect_i2 - sect_i1, sect_j2 - sect_j1)):
                            old_line_val: str | None = (
                                old_section[sect_i1 + line_idx]
                                if sect_i1 + line_idx < len(old_section)
                                else None
                            )
                            new_line_val: str | None = (
                                new_section[sect_j1 + line_idx]
                                if sect_j1 + line_idx < len(new_section)
                                else None
                            )

                            if old_line_val and new_line_val:
                                # both exist - ALWAYS use word-level diff to show only changed words
                                # this ensures that adding/removing a single word only highlights that word
                                old_parts, new_parts = _compute_char_diff(
                                    old_line_val, new_line_val, word_level=True
                                )

                                has_removed = any(
                                    part.get("type") == "removed" for part in old_parts
                                )
                                has_added = any(part.get("type") == "added" for part in new_parts)

                                if has_added and not has_removed:
                                    change_type = "addition_only"
                                elif has_removed and not has_added:
                                    change_type = "removal_only"
                                else:
                                    change_type = "mixed"

                                diff_segments.append(
                                    {
                                        "type": "modified",  # line exists in both versions - use word-level diff
                                        "change_type": change_type,
                                        "old_start": i1 + sect_i1 + line_idx + 1,
                                        "old_end": i1 + sect_i1 + line_idx + 1,
                                        "new_start": j1 + sect_j1 + line_idx + 1,
                                        "new_end": j1 + sect_j1 + line_idx + 1,
                                        "old_lines": [old_line_val],
                                        "new_lines": [new_line_val],
                                        "old_char_diffs": [old_parts],  # word-level diff parts
                                        "new_char_diffs": [new_parts],  # word-level diff parts
                                    }
                                )
                            elif old_line_val:
                                # old line exists but new doesn't - removed
                                diff_segments.append(
                                    {
                                        "type": "removed",
                                        "old_start": i1 + sect_i1 + line_idx + 1,
                                        "old_end": i1 + sect_i1 + line_idx + 1,
                                        "old_lines": [old_line_val],
                                        "new_lines": [],
                                    }
                                )
                            elif new_line_val:
                                # new line exists but old doesn't - added
                                diff_segments.append(
                                    {
                                        "type": "added",
                                        "new_start": j1 + sect_j1 + line_idx + 1,
                                        "new_end": j1 + sect_j1 + line_idx + 1,
                                        "old_lines": [],
                                        "new_lines": [new_line_val],
                                    }
                                )

    return diff_segments


# generate human-readable summary of file changes for the diff viewer
def _nl_summary_for_diff(
    path: str,
    approx_bytes_changed: int,
    percent: float,
    ranges: list[tuple[int, int]],
    zip_changes: list[dict[str, Any]],
    chunk_size: int,
    base_size: int,
    cur_size: int,
) -> dict[str, str]:
    """Return {'headline': str, 'details': str}"""
    kind = _nl_file_kind(path)
    # headline
    if approx_bytes_changed == 0 and not zip_changes:
        headline = f"No differences detected in the {kind}."
        return {"headline": headline, "details": ""}

    # zip member changes summary
    added = [z for z in zip_changes if z.get("change") == "added"]
    removed = [z for z in zip_changes if z.get("change") == "removed"]
    modified = [z for z in zip_changes if z.get("change") == "modified"]

    chunk_regions = len(ranges)
    headline = f"{kind.capitalize()} changed ~{percent:.2f}% ({approx_bytes_changed} bytes)."

    parts: list[str] = []
    # file sizes
    if base_size and cur_size and base_size != cur_size:
        delta = cur_size - base_size
        sign = "+" if delta >= 0 else ""
        parts.append(f"Size: {base_size} → {cur_size} bytes ({sign}{delta}).")

    # regions by chunks
    if chunk_regions:
        total_chunks = 0
        for s, e in ranges:
            total_chunks += e - s + 1
        parts.append(
            f"Changed regions: {chunk_regions} ({total_chunks} chunk(s) of {chunk_size} bytes)."
        )

    # Office containers (docx/xlsx/pptx)
    if zip_changes:
        a, r, m = len(added), len(removed), len(modified)
        bits = []
        if a:
            bits.append(f"{a} added")
        if r:
            bits.append(f"{r} removed")
        if m:
            bits.append(f"{m} modified")
        parts.append("ZIP members: " + ", ".join(bits) + ".")

    # heuristics by type
    if kind in ("PowerPoint", "Word document", "Excel workbook"):
        if zip_changes and modified:
            parts.append("Likely content edits to internal parts (slides, document XML, or media).")
        elif zip_changes and (added or removed):
            parts.append("Likely added or removed embedded media, slides, or sheets.")
        else:
            parts.append("Container changed; content may have been edited or metadata updated.")
    elif kind == "PDF":
        parts.append("Binary regions changed; could be page content or metadata.")
    elif kind == "source code":
        parts.append("Code regions changed; review in your VCS for exact lines.")
    elif kind == "text file":
        parts.append("Text content changed; open in a diff tool for exact lines.")

    details = " ".join(parts)
    return {"headline": headline, "details": details}


# -------- Baseline snapshot storage (content-addressed) --------
# optional full-file snapshot storage: stores complete file content in a content-addressed
# storage system (same content = same hash = same file, deduplication for free)
# automatically prunes old snapshots to stay within size limits
BASELINES_DIR = str((BASE_DIR / "data" / "baselines").resolve())
BASELINES_MAX_BYTES: int = int(getattr(CFG, "baselines_max_bytes", 1_000_000_000))  # ~1 GB default


# shard directories by hash prefix to avoid huge flat directories
def _shard_dir_from_hex(h: str) -> str:
    # shard to avoid huge directories
    return os.path.join(BASELINES_DIR, h[:2], h[2:4], h)


# store a full file snapshot in content-addressed storage (same content = same hash = same file)
def _snapshot_file_to_cas(src_path: str) -> dict[str, Any]:
    """
    save the *bytes* of src_path into content-addressed storage by SHA-256.
    returns a metadata dict suitable for storing in target["baseline_blob"].
    """
    src = _norm_user_path(src_path)
    st = os.stat(src)
    size = int(st.st_size)

    h = hashlib.sha256()
    with open(src, "rb") as f:
        for chunk in iter(lambda: f.read(1024 * 1024), b""):
            h.update(chunk)
    sha = h.hexdigest().upper()

    dst_root = _shard_dir_from_hex(sha)
    os.makedirs(dst_root, exist_ok=True)

    fname = os.path.basename(src_path) or sha
    dst_path = os.path.join(dst_root, fname)

    if not os.path.exists(dst_path):
        tmp = dst_path + ".tmp"
        with open(src, "rb") as rf, open(tmp, "wb") as wf:
            for chunk in iter(lambda: rf.read(1024 * 1024), b""):
                wf.write(chunk)
        os.replace(tmp, dst_path)

    return {
        "sha256": sha,
        "size": size,
        "stored_path": dst_path,
        "created_at": int(time.time()),
    }


# validate that a stored baseline blob is still intact (hash matches stored path)
def _validate_cas_blob(blob: dict[str, Any]) -> bool:
    try:
        p = blob.get("stored_path")
        if not p or not os.path.exists(p):
            return False
        if int(os.path.getsize(p)) != int(blob.get("size", -1)):
            return False
        # optional: verify hash matches
        h = hashlib.sha256()
        with open(p, "rb") as f:
            for chunk in iter(lambda: f.read(1024 * 1024), b""):
                h.update(chunk)
        return h.hexdigest().upper() == str(blob.get("sha256", "")).upper()
    except Exception:
        return False


# prune old baseline snapshots to stay within size limits (oldest files first)
def _maybe_prune_baselines() -> None:
    """
    keep total size of BASELINES_DIR under BASELINES_MAX_BYTES.
    deletes oldest files first. called after adding new baseline blobs.
    """
    try:
        files: list[tuple[str, float, int]] = []
        total_size = 0

        # walk recursively and collect all baseline files
        for root, _dirs, fnames in os.walk(BASELINES_DIR):
            for fn in fnames:
                p = os.path.join(root, fn)
                try:
                    st = os.stat(p)
                    total_size += st.st_size
                    files.append((p, st.st_mtime, st.st_size))
                except Exception:
                    continue

        # if we are over the cap, remove oldest files until under limit
        if total_size > BASELINES_MAX_BYTES:
            files.sort(key=lambda x: x[1])  # oldest first
            for p, _mtime, sz in files:
                try:
                    os.remove(p)
                    total_size -= sz
                    if total_size <= BASELINES_MAX_BYTES:
                        break
                except Exception:
                    pass
    except Exception:
        pass


# ---------------- Flask app build ----------------
# build the Flask application with all routes and event processing logic
def build_app(event_bus) -> Flask:
    app = Flask(__name__)

    # configure session security (httpOnly, sameSite, secure when HTTPS)
    # sessions are NOT persistent - cookies expire when browser closes (session-only)
    # combine the base secret with a per-run session key to invalidate old sessions on program restart
    # this ensures users must login again each time the program launches
    # SESSION_SECRET is guaranteed to be str after validation in auth.py
    assert SESSION_SECRET is not None
    app.secret_key = SESSION_SECRET + _current_session_key
    app.config["SESSION_COOKIE_HTTPONLY"] = True  # prevent XSS attacks
    app.config["SESSION_COOKIE_SAMESITE"] = "Lax"  # CSRF protection
    # secure cookies only over HTTPS (default to False for localhost, set to True in production with HTTPS)
    # check if we're running on HTTPS by examining the host config or environment
    app.config["SESSION_COOKIE_SECURE"] = os.getenv("CUSTOSEYE_HTTPS", "false").lower() in (
        "true",
        "1",
        "yes",
    )
    # session cookies expire when browser closes (non-persistent sessions)
    # do not set PERMANENT_SESSION_LIFETIME - sessions will be session-only by default
    app.config["SESSION_COOKIE_NAME"] = "custoseye_session"  # custom session cookie name

    # make sure our iterator type is clear for mypy
    _iter: Iterator[dict[str, Any]] = _bus_iterator(event_bus)

    # publisher helper used to push live events immediately (for integrity events from API calls)
    def _publish(ev: dict[str, Any]) -> None:
        try:
            if hasattr(event_bus, "publish"):
                event_bus.publish(ev)  # preferred
            else:
                # fallback: push into BUFFER so it appears right away
                BUFFER.append(ev)
        except Exception:
            # never break the request on failed publish
            pass

    # emit an integrity event to the live event stream (for manual hash operations)
    def _emit_integrity_event(path: str, status: str, rule: str) -> None:
        """
        push an integrity-related event into the live event stream.
        used by /api/integrity/hash so that users see CHANGED/OK immediately.
        skips emitting events for OK statuses (baseline updated/deletion accepted)
        since we update the existing CRITICAL entry instead.
        """
        s = (status or "").upper()
        # skip emitting new events for OK statuses, we update existing CRITICAL entries instead
        if s.startswith("OK") and (
            "baseline updated" in status.lower() or "deletion accepted" in status.lower()
        ):
            return

        lvl = "info"
        if s.startswith("CHANGED"):
            lvl = "critical"
        elif s.startswith("ERR"):
            lvl = "warning"

        ev = {
            "source": "integrity",
            "level": lvl,
            "reason": f"integrity {status or 'update'}",
            "path": path,
            "rule": rule,
            "ts": time.time(),
            "name": os.path.basename(path) if path else "",
        }
        _publish(ev)

    # static asset routes (favicon and touch icons for web UI)
    @app.get("/favicon.ico")
    def favicon():
        return app.send_static_file("assets/favicon.ico")

    @app.get("/favicon-32x32.png")
    def favicon_32():
        return app.send_static_file("assets/favicon-32x32.png")

    @app.get("/favicon-16x16.png")
    def favicon_16():
        return app.send_static_file("assets/favicon-16x16.png")

    @app.get("/apple-touch-icon.png")
    def apple_touch_icon():
        return app.send_static_file("assets/apple-touch-icon.png")

    # main event processing pipeline: drain events from bus, process them, and add to buffer
    def drain_into_buffer() -> int:
        with _DRAIN_LOCK:
            _maybe_reload_rules()
            _maybe_reload_name_trust()
            drained = 0
            deadline = time.time() + DRAIN_DEADLINE_SEC

            while time.time() < deadline and drained < DRAIN_LIMIT_PER_CALL:
                try:
                    ev = next(_iter)
                except StopIteration:
                    break
                except Exception:
                    break
                if ev is None:
                    continue  # do not exit early, keep trying until deadline

                ev.setdefault("level", "info")
                ev.setdefault("reason", "event")
                ev.setdefault("ts", time.time())

                if ev.get("source") == "integrity":
                    # update integrity targets when integrity events come in
                    _update_integrity_target_from_event(ev)

                    # check if this is a "Hash verified" event that should update existing CRITICAL entry
                    if ev.get("update_existing") and ev.get("reason") == "Hash verified":
                        # update existing CRITICAL entry instead of creating new one
                        path = ev.get("path", "")
                        if path:
                            _update_live_events_for_hash_verified(path)
                            # do not add this event to buffer, we updated the existing entry
                            continue
                else:
                    # makes sure we always have a dict for mypy and runtime safety
                    decision: dict[str, Any] = cast(dict[str, Any], _rules.evaluate(ev) or {})
                    ev["level"] = decision.get("level", ev.get("level"))
                    ev["reason"] = decision.get("reason", ev.get("reason"))

                if ev.get("source") == "process":
                    # --- suppress our own process(es) before any scoring or tree updates ---
                    nm_lc = str(ev.get("name") or "").lower()
                    ex_lc = str(ev.get("exe") or "").lower()
                    h_lc = str(ev.get("sha256") or "").lower()
                    if (
                        (nm_lc and nm_lc in SELF_SUPPRESS["names"])
                        or (ex_lc and ex_lc in SELF_SUPPRESS["exes"])
                        or (h_lc and h_lc in SELF_SUPPRESS["sha256"])
                    ):
                        continue  # skip buffering and tree indexing entirely

                    # --- trust scoring: name-based fast-path, then kernel/idle/registry, then model ---
                    pid = ev.get("pid")
                    nm = str(ev.get("name") or "")

                    # 1) known-good names (your NAME_TRUST map)
                    if _maybe_promote_to_trusted(ev):
                        pass  # already set by the heuristic

                    # 2) kernel/idle/registry fast-path
                    elif pid in (0, 4) or nm.lower() in (
                        "system",
                        "system idle process",
                        "registry",
                    ):
                        ev["csc"] = {
                            "version": "v2",
                            "verdict": "trusted",
                            "cls": "system",
                            "confidence": 0.98,
                            "reasons": ["kernel/idle/registry fast-path"],
                            "signals": {},
                        }
                        ev["csc_verdict"] = "trusted"
                        ev["csc_class"] = "system"
                        ev["csc_confidence"] = 0.98
                        ev["csc_reasons"] = ["kernel/idle/registry fast-path"]

                    # 3) everything else -> model
                    else:
                        csc_out = _csc.evaluate(ev)
                        ev["csc"] = csc_out
                        ev["csc_verdict"] = csc_out.get("verdict", "unknown")
                        ev["csc_class"] = csc_out.get("cls", "unknown")
                        ev["csc_confidence"] = float(csc_out.get("confidence", 0.5))
                        ev["csc_reasons"] = csc_out.get("reasons", [])

                    # process tree index (store the v2 fields)
                    if isinstance(pid, int):
                        PROC_INDEX[pid] = {
                            "pid": pid,
                            "ppid": ev.get("ppid"),
                            "name": ev.get("name") or "",
                            "csc_verdict": ev.get("csc_verdict", "unknown"),
                            "csc_class": ev.get("csc_class", "unknown"),
                            "csc_confidence": ev.get("csc_confidence", 0.5),
                        }

                if ev.get("source") == "network" and not (ev.get("pid") or ev.get("name")):
                    continue

                # --- integrity name fill ---
                if ev.get("source") == "integrity":
                    p = ev.get("path") or ev.get("file")
                    if p and not ev.get("name"):
                        try:
                            ev["name"] = os.path.basename(str(p))
                        except Exception:
                            ev["name"] = str(p)

                # --- safe coalesce (don’t drop worse events) ---
                if _coalesce_or_admit(ev):
                    BUFFER.append(ev)
                    drained += 1
                # else: merged into a recent identical event; no append

            return drained

    # register authentication routes (must be registered before other routes)
    register_auth_routes(app)

    # preview route: serve dashboard for background preview (no auth required, read-only)
    @app.get("/preview")
    def preview():
        """serve dashboard preview for login page background - read-only, non-interactive"""
        return render_template("index.html")

    # preview API endpoints: allow preview mode to fetch data without auth (read-only)
    @app.get("/api/preview/events")
    def preview_events():
        """preview endpoint for events - no auth required"""
        drain_into_buffer()
        include_info = (request.args.get("include_info") or "").lower() in ("1", "true", "yes")
        data = []
        for ev in BUFFER:
            if ev.get("source") == "network" and not (ev.get("pid") or ev.get("name")):
                continue
            lvl = (ev.get("level") or "info").lower()
            if not include_info and lvl == "info":
                continue
            data.append(ev)
        return jsonify(data)

    @app.get("/api/preview/integrity/targets")
    def preview_integrity_targets():
        """preview endpoint for integrity targets - no auth required"""
        return jsonify(_read_integrity_targets())

    @app.get("/api/preview/about")
    def preview_about():
        """preview endpoint for about info - no auth required"""
        return jsonify({"version": "preview"})

    @app.get("/api/preview/proctree")
    def preview_proctree():
        """preview endpoint for process tree - no auth required"""
        drain_into_buffer()
        children: dict[int, list[int]] = {}
        roots: list[int] = []
        for pid, rec in PROC_INDEX.items():
            ppid = rec.get("ppid")
            if isinstance(ppid, int) and ppid in PROC_INDEX:
                children.setdefault(ppid, []).append(pid)
            else:
                roots.append(pid)

        def build(pid: int) -> dict[str, Any]:
            r = PROC_INDEX.get(pid, {})
            return {
                "pid": pid,
                "ppid": r.get("ppid", None),
                "name": r.get("name", ""),
                "csc_verdict": r.get("csc_verdict", "unknown"),
                "csc_class": r.get("csc_class", "unknown"),
                "csc_confidence": float(r.get("csc_confidence", 0.5)),
                "children": [build(c) for c in children.get(pid, [])],
            }

        tree = [build(r) for r in sorted(roots)]
        if request.args.get("as") == "json":
            response = make_response(json.dumps(tree, indent=2))
            response.headers["Content-Type"] = "application/json"
            response.headers["Content-Disposition"] = (
                'attachment; filename="custoseye_proctree.json"'
            )
            return response
        return jsonify(tree)

    # main page route: serve the dashboard HTML (requires authentication)
    @app.get("/")
    @require_auth
    def index():
        return render_template("index.html")

    # event stream API: return events from the buffer (with optional filtering)
    @app.get("/api/events")
    @require_auth
    def events():
        drain_into_buffer()
        include_info = (request.args.get("include_info") or "").lower() in ("1", "true", "yes")
        data = []
        for ev in BUFFER:
            if ev.get("source") == "network" and not (ev.get("pid") or ev.get("name")):
                continue
            lvl = (ev.get("level") or "info").lower()
            if not include_info and lvl == "info":
                continue
            data.append(ev)
        return jsonify(data)

    # ping endpoint: lightweight drain trigger to keep event ingestion going
    @app.get("/api/ping")
    @require_auth
    def ping():
        """lightweight drain trigger so background ingestion continues on any tab"""
        n = drain_into_buffer()
        return jsonify({"ok": True, "drained": n, "buffer": len(BUFFER)})

    # export endpoint: export events from buffer in various formats (CSV, JSON, JSONL, XLSX)
    @app.get("/api/export")
    @require_auth
    def export_current():
        """
        export live buffer with current filters.
        format=csv (default) | json | jsonl | xlsx
        """
        drain_into_buffer()

        include_info = (request.args.get("include_info") or "").lower() in ("1", "true", "yes")
        q = (request.args.get("q") or "").lower().strip()
        lvls = [x for x in (request.args.get("levels") or "").lower().split(",") if x]
        fmt = (request.args.get("format") or "csv").lower()

        def pass_filters(ev: dict[str, Any]) -> bool:
            lvl = (ev.get("level") or "info").lower()
            if lvl == "info" and not include_info:
                return False
            if lvls and lvl not in lvls:
                return False
            if q:
                s = f"{ev.get('reason','')} {ev.get('source','')} {ev.get('name','')} {ev.get('pid','')} {ev.get('path','')}"
                if q not in s.lower():
                    return False
            return True

        rows = [ev for ev in BUFFER if pass_filters(ev)]

        # JSONL
        if fmt == "jsonl":
            lines = [json.dumps(ev, ensure_ascii=False) for ev in rows]
            resp = make_response("\n".join(lines))
            resp.headers["Content-Type"] = "application/x-ndjson"
            resp.headers["Content-Disposition"] = 'attachment; filename="custoseye_export.jsonl"'
            return resp

        # JSON
        if fmt == "json":
            resp = make_response(json.dumps(rows, ensure_ascii=False, indent=2))
            resp.headers["Content-Type"] = "application/json"
            resp.headers["Content-Disposition"] = 'attachment; filename="custoseye_export.json"'
            return resp

        # common tabular cols
        cols = [
            "ts",
            "level",
            "reason",
            "source",
            "pid",
            "name",
            "path",
            "csc_verdict",
            "csc_class",
            "csc_confidence",
        ]

        # XLSX
        if fmt == "xlsx":
            try:
                from openpyxl import Workbook
                from openpyxl.utils import get_column_letter
            except Exception:
                return (
                    jsonify(
                        {
                            "error": "xlsx export requires `openpyxl`",
                            "hint": "pip install openpyxl or use format=csv",
                        }
                    ),
                    400,
                )

            wb = Workbook()
            ws = wb.active
            ws.title = "CustosEye"
            ws.append(cols)

            for ev in rows:
                r = {k: ev.get(k, "") for k in cols}
                if isinstance(r["ts"], int | float):
                    try:
                        r["ts"] = time.strftime("%Y-%m-%d %H:%M:%S", time.localtime(float(r["ts"])))
                    except Exception:
                        pass
                ws.append([r.get(c, "") for c in cols])

            for i, c in enumerate(cols, 1):
                max_len = len(c)
                for row in ws.iter_rows(min_row=2, min_col=i, max_col=i):
                    v = row[0].value
                    if v is None:
                        continue
                    max_len = max(max_len, len(str(v)))
                ws.column_dimensions[get_column_letter(i)].width = max(
                    10, min(60, int(max_len * 1.1 + 2))
                )

            from io import BytesIO

            bio = BytesIO()
            wb.save(bio)
            bio.seek(0)
            return send_file(
                bio,
                as_attachment=True,
                download_name="custoseye_export.xlsx",
                mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )

        # default: CSV (Excel-friendly, BOM + quoted + CRLF; ts as text)
        buf = io.StringIO(newline="")
        writer = csv.DictWriter(
            buf,
            fieldnames=cols,
            extrasaction="ignore",
            quoting=csv.QUOTE_ALL,
            lineterminator="\r\n",
        )
        writer.writeheader()

        for ev in rows:
            r = {k: ev.get(k, "") for k in cols}
            if isinstance(r["ts"], int | float):
                try:
                    r["ts"] = time.strftime("%Y-%m-%d %H:%M:%S", time.localtime(float(r["ts"])))
                except Exception:
                    pass
            if r.get("ts"):
                r["ts"] = f"'{r['ts']}"
            writer.writerow(r)

        csv_text = buf.getvalue()
        out_bytes = ("\ufeff" + csv_text).encode("utf-8")
        resp = make_response(out_bytes)
        resp.headers["Content-Type"] = "text/csv; charset=utf-8"
        resp.headers["Content-Disposition"] = 'attachment; filename="custoseye_export.csv"'
        return resp

    # process tree API: return the process tree built from process events
    @app.get("/api/proctree")
    @require_auth
    def proctree():
        """
        return the compact process tree.

        query params:
          as=json  -> download pretty JSON (custoseye_proctree.json)
          (none)   -> return JSON to the browser (not as attachment)
        """
        drain_into_buffer()

        children: dict[int, list[int]] = {}
        roots: list[int] = []
        for pid, rec in PROC_INDEX.items():
            ppid = rec.get("ppid")
            if isinstance(ppid, int) and ppid in PROC_INDEX:
                children.setdefault(ppid, []).append(pid)
            else:
                roots.append(pid)

        def build(pid: int) -> dict[str, Any]:
            r = PROC_INDEX.get(pid, {})
            return {
                "pid": pid,
                "ppid": r.get("ppid", None),
                "name": r.get("name", ""),
                "csc_verdict": r.get("csc_verdict", "unknown"),
                "csc_class": r.get("csc_class", "unknown"),
                "csc_confidence": float(r.get("csc_confidence", 0.5)),
                "children": [
                    build(c) for c in sorted(children.get(pid, []))[: CFG.max_tree_children]
                ],
            }

        tree = [build(p) for p in sorted(roots)[: CFG.max_tree_roots]]

        fmt = (request.args.get("as") or "").lower()
        if fmt == "json":
            resp = make_response(json.dumps(tree, indent=2, ensure_ascii=False))
            resp.headers["Content-Type"] = "application/json"
            resp.headers["Content-Disposition"] = 'attachment; filename="custoseye_proctree.json"'
            return resp

        # default inline JSON
        return jsonify(tree)

    # about endpoint: return version and build info from VERSION.txt
    @app.get("/api/about")
    @require_auth
    def about():
        version = "-"
        build = "-"
        vpath = BASE_DIR / "VERSION.txt"
        if vpath.exists():
            try:
                lines = [
                    line.strip()
                    for line in vpath.read_text(encoding="utf-8").splitlines()
                    if line.strip()
                ]
                if len(lines) >= 1:
                    version = lines[0]
                if len(lines) >= 2:
                    build = lines[1]
            except Exception:
                pass
        return jsonify({"version": version, "build": build, "buffer_max": BUFFER_MAX})

        # ---------------- integrity endpoints ----------------

    # integrity targets API: get the watch list
    @app.get("/api/integrity/targets")
    @require_auth
    def api_integrity_targets_get():
        # return normalized list
        return jsonify(_read_integrity_targets())

    # integrity targets API: add or update a file on the watch list
    @app.post("/api/integrity/targets")
    @require_auth
    def api_integrity_targets_post():
        """
        body: { path, rule, note? }
        rule: "sha256" | "mtime+size"
        behavior: if rule == "sha256" and no baseline stored yet, auto-hash now to set baseline.
        """
        try:
            body: dict[str, Any] = cast(dict[str, Any], request.get_json(force=True) or {})
            path = str(body.get("path") or "").strip()
            rule = str(body.get("rule") or "sha256").strip().lower()
            note = str(body.get("note") or "").strip()
            if not path:
                return jsonify({"ok": False, "error": "missing path"}), 400
            if rule not in ("sha256", "mtime+size"):
                return jsonify({"ok": False, "error": "invalid rule"}), 400

            rows = _read_integrity_targets()

            # update or insert
            target = None
            for r in rows:
                if str(r.get("path") or "").lower() == path.lower():
                    target = r
                    break
            if target is None:
                target = {"path": path, "rule": rule}
                if note != "":
                    target["note"] = note
                else:
                    # allow clearing the note explicitly with empty string
                    if "note" in target:
                        del target["note"]
                rows.append(target)
            else:
                target["path"] = path  # keep users original text (env vars allowed)
                target["rule"] = rule
                if note != "":
                    target["note"] = note
                else:
                    # allow clearing the note explicitly with empty string
                    if "note" in target:
                        del target["note"]

            # auto-baseline for sha256 if missing
            if rule == "sha256" and not target.get("sha256"):
                info = _sha256_file(path)
                if info.get("sha256"):
                    target["sha256"] = info["sha256"]
                    target["last_result"] = "OK (baseline set)"
                    if ALLOW_MTIME_SNAPSHOTS:
                        try:
                            target["baseline_blob"] = _snapshot_file_to_cas(path)
                            _maybe_prune_baselines()
                        except Exception as e:
                            target["baseline_blob"] = {"error": str(e)}
                else:
                    # leave baseline empty; UI can re-hash later
                    target["last_result"] = f"ERR: {info.get('error', 'hash failed')}"

            # auto-baseline for mtime+size if missing
            if rule == "mtime+size" and (not target.get("mtime") or not target.get("size")):
                try:
                    st = os.stat(_norm_user_path(path))
                    target["mtime"] = int(st.st_mtime)
                    target["size"] = int(st.st_size)
                    target["last_result"] = "OK (baseline set)"
                    if ALLOW_MTIME_SNAPSHOTS:
                        try:
                            target["baseline_blob"] = _snapshot_file_to_cas(path)
                            _maybe_prune_baselines()
                        except Exception as e:
                            target["baseline_blob"] = {"error": str(e)}
                except Exception as e:
                    target["last_result"] = f"ERR: {e}"

            # build / refresh chunk baseline for diffs (both rules benefit)
            # note: we don't persist file content, only per-chunk hashes
            info_chunks = _chunk_hashes(path)
            if "error" not in info_chunks:
                target["chunks"] = info_chunks  # algo, chunk_size, size, hashes
            else:
                # leave chunks absent if we couldn't read; UI can re-baseline later
                target.pop("chunks", None)

            # If it is a ZIP (docx, xlsx, etc), store a member manifest for friendlier diffs
            zman = _zip_manifest(path)
            if zman:
                target["zip_manifest"] = zman
            else:
                target.pop("zip_manifest", None)

            _write_integrity_targets(rows)
            return jsonify({"ok": True})

        except Exception as e:
            return jsonify({"ok": False, "error": str(e)}), 500

    # integrity targets API: remove a file from the watch list
    @app.delete("/api/integrity/targets")
    @require_auth
    def api_integrity_targets_delete():
        try:
            body: dict[str, Any] = cast(dict[str, Any], request.get_json(force=True) or {})
            path = str(body.get("path") or "").strip()
            if not path:
                return jsonify({"ok": False, "error": "missing path"}), 400
            rows = _read_integrity_targets()

            # Check if file was actually in the watch list before removing
            was_in_list = any(str(r.get("path") or "").lower() == path.lower() for r in rows)

            # Remove the file from the watch list
            rows2 = [r for r in rows if str(r.get("path") or "").lower() != path.lower()]
            _write_integrity_targets(rows2)

            # Emit a live event to notify that the file was removed from the watch list
            if was_in_list:
                ev = {
                    "source": "integrity",
                    "level": "critical",
                    "reason": f"File removed from watch list: {path}",
                    "path": path,
                    "ts": time.time(),
                    "name": os.path.basename(path) if path else "",
                }
                _publish(ev)

            return jsonify({"ok": True})
        except Exception as e:
            return jsonify({"ok": False, "error": str(e)}), 500

    # integrity hash API: manually rehash a file and update its baseline
    @app.post("/api/integrity/hash")
    @require_auth
    def api_integrity_hash():
        """
        re-hash: manually compute SHA-256 and update baseline if needed.
        this is for manually generating a new trusted hash, NOT for detection.
        does NOT emit events to avoid duplicate alerts, the integrity checker handles detection.
        """
        try:
            body: dict[str, Any] = cast(
                dict[str, Any], request.get_json(force=True) or {}
            )  # parse JSON body
            path = str(body.get("path") or "").strip()  # normalize to a simple string
            if not path:
                return jsonify({"error": "missing path"}), 400  # hard fail if no path

            # check if file exists first, handle deletion gracefully
            if not os.path.exists(_norm_user_path(path)):
                # file is deleted, update status gracefully
                rows = _read_integrity_targets()
                for r in rows:
                    if str(r.get("path") or "").lower() == path.lower():
                        r["last_result"] = "CHANGED: File deleted or path changed"
                        _write_integrity_targets(rows)
                        return jsonify(
                            {
                                "error": "File deleted or missing",
                                "last_result": "CHANGED: File deleted or path changed",
                                "file_deleted": True,
                            }
                        )

            info = _sha256_file(path)  # compute current file hash and attrs
            rows = _read_integrity_targets()  # load current targets file
            changed = False  # track if we modified the stored record

            # update corresponding row if present
            for r in rows:
                if (
                    str(r.get("path") or "").lower() == path.lower()
                ):  # match on case-insensitive path
                    rule = str(r.get("rule") or "sha256").lower()  # either "sha256" or "mtime+size"
                    baseline = str(r.get("sha256") or "")  # baseline hash if present

                    if "error" in info:
                        # hashing failed, record the error but do not change baselines
                        r["last_result"] = f"ERR: {info['error']}"
                        changed = True
                        # DO NOT emit event, this is manual Re-hash, not detection
                    else:
                        new_hash = info.get("sha256", "")  # current SHA-256 hex
                        if rule == "sha256":
                            if not baseline:
                                # first time, promote to baseline
                                r["sha256"] = new_hash
                                r["last_result"] = "OK (baseline set)"
                                changed = True
                                # DO NOT emit event, this is manual Re-hash, not detection
                                if ALLOW_MTIME_SNAPSHOTS:
                                    try:
                                        r["baseline_blob"] = _snapshot_file_to_cas(path)
                                        _maybe_prune_baselines()
                                    except Exception as e:
                                        r["baseline_blob"] = {"error": str(e)}
                            else:
                                # simple equality check against baseline
                                r["last_result"] = "OK" if baseline == new_hash else "CHANGED"
                                changed = True
                                # DO NOT emit event, this is manual Re-hash, not detection
                        else:
                            # mtime+size rule, compare against or establish baseline
                            try:
                                st = os.stat(_norm_user_path(path))
                                cur_mtime = int(st.st_mtime)
                                cur_size = int(st.st_size)
                                if not r.get("mtime") or not r.get("size"):
                                    # no baseline yet, establish it now
                                    r["mtime"], r["size"] = cur_mtime, cur_size
                                    r["last_result"] = "OK (baseline set)"
                                    changed = True
                                    _emit_integrity_event(path, r["last_result"], rule)
                                    if ALLOW_MTIME_SNAPSHOTS:
                                        try:
                                            r["baseline_blob"] = _snapshot_file_to_cas(path)
                                            _maybe_prune_baselines()
                                        except Exception as e:
                                            r["baseline_blob"] = {"error": str(e)}
                                else:
                                    # compare to existing baseline
                                    r["last_result"] = (
                                        "OK"
                                        if (r["mtime"] == cur_mtime and r["size"] == cur_size)
                                        else "CHANGED"
                                    )
                                    changed = True
                                    # DO NOT emit event, this is manual Re-hash, not detection
                            except Exception as e:
                                r["last_result"] = f"ERR: {e}"
                                # DO NOT emit event, this is manual Re-hash, not detection
                                changed = True

                    # make sure we have per chunk baseline so the diff endpoint can work
                    # we only store per chunk hashes, not content
                    if "error" not in info:
                        need_chunks = "chunks" not in r or not isinstance(
                            r.get("chunks", {}).get("hashes"), list
                        )
                        if need_chunks:
                            ch = _chunk_hashes(path)  # build per chunk baseline
                            if "error" not in ch:
                                r["chunks"] = ch  # algo, chunk_size, size, hashes
                                changed = True

                        # only capture zip_manifest when (and only when) we establish a baseline
                        if "error" not in info:
                            if (rule == "sha256" and not baseline) or (
                                rule == "mtime+size" and (not r.get("mtime") or not r.get("size"))
                            ):
                                zman = _zip_manifest(path)
                                if zman:
                                    r["zip_manifest"] = zman
                                    changed = True

                    break  # stop after the first matching record

            # if we edited anything, persist the targets file
            if changed:
                _write_integrity_targets(rows)
            if changed:
                # find the updated row again to read last_result
                for r in rows:
                    if str(r.get("path") or "").lower() == path.lower():
                        info["last_result"] = r.get("last_result", "")
                        info["rule"] = r.get("rule", "")
                        break

            # return raw info for the alert preview in UI
            return jsonify(info)

        except Exception as e:
            return jsonify({"error": str(e)}), 500

    # integrity diff API: compute privacy-preserving diff between baseline and current file
    @app.post("/api/integrity/diff")
    @require_auth
    def api_integrity_diff():
        """
        privacy-preserving diff between baseline chunk hashes and the current file.

        we never return file contents, only:
        - which chunk ranges changed (by chunk-hash),
        - tiny "after" previews (hex + ASCII) capped per region.

        percent logic:
        - for ZIP-based Office files (docx/xlsx/pptx), estimate change from ZIP
            member deltas (added/removed = size, modified = abs(size diff), or a
            small cap when only CRC differs). Ignore the chunk-floor here to avoid
            100% spikes from container re-packing.
        - for non-ZIP files, fall back to chunk-region coverage as before.
        """
        try:
            body: dict[str, Any] = cast(dict[str, Any], request.get_json(force=True) or {})
            path = str(body.get("path") or "").strip()
            max_regions = int(body.get("max_regions") or 50)
            preview_cap = int(body.get("preview_bytes") or 64)

            if not path:
                return jsonify({"ok": False, "error": "missing path"}), 400

            # find target + baseline chunks
            rows = _read_integrity_targets()
            target = next(
                (r for r in rows if str(r.get("path") or "").lower() == path.lower()), None
            )
            if target is None:
                return jsonify({"ok": False, "error": "not on watch list"}), 400

            baseline = target.get("chunks")
            if not baseline or not isinstance(baseline.get("hashes"), list):
                return jsonify({"ok": False, "error": "no baseline chunks available"}), 400

            # check if file exists
            if not os.path.exists(_norm_user_path(path)):
                return (
                    jsonify(
                        {"ok": False, "error": "File deleted or path changed", "file_deleted": True}
                    ),
                    400,
                )

            # current per-chunk hashes
            cur = _chunk_hashes(path)
            if "error" in cur:
                return jsonify({"ok": False, "error": cur["error"]}), 400

            base_hashes = list(baseline.get("hashes") or [])
            cur_hashes = list(cur.get("hashes") or [])
            chunk_size = int(cur.get("chunk_size") or baseline.get("chunk_size") or 4096)
            base_size = int(baseline.get("size") or 0)
            cur_size = int(cur.get("size") or 0)

            # ----- ZIP member-level comparison (for docx/xlsx/pptx) -----
            zip_before = target.get("zip_manifest") or {}
            zip_after = _zip_manifest(path)
            zip_changes: list[dict[str, Any]] = []

            if zip_after or zip_before:
                before_keys = set(zip_before.keys())
                after_keys = set(zip_after.keys())
                added = sorted(after_keys - before_keys)
                removed = sorted(before_keys - after_keys)
                common = sorted(before_keys & after_keys)

                for k in added:
                    z = zip_after[k]
                    zip_changes.append(
                        {"member": k, "change": "added", "size": z["size"], "crc": z["crc"]}
                    )
                for k in removed:
                    z = zip_before[k]
                    zip_changes.append(
                        {"member": k, "change": "removed", "size": z["size"], "crc": z["crc"]}
                    )
                for k in common:
                    a, b = zip_after[k], zip_before[k]
                    if a["size"] != b["size"] or a["crc"] != b["crc"]:
                        zip_changes.append(
                            {
                                "member": k,
                                "change": "modified",
                                "size_before": b["size"],
                                "size_after": a["size"],
                                "crc_before": b["crc"],
                                "crc_after": a["crc"],
                            }
                        )

            # ----- identify changed chunk indices (for regions UI only) -----
            max_len = max(len(base_hashes), len(cur_hashes))
            changed_idxs: list[int] = []
            for i in range(max_len):
                old = base_hashes[i] if i < len(base_hashes) else None
                new = cur_hashes[i] if i < len(cur_hashes) else None
                if old != new:
                    changed_idxs.append(i)

            ranges = _merge_changed_chunks(changed_idxs)

            # build regions with offsets and tiny "after" previews
            regions_out: list[dict[str, Any]] = []
            chunk_floor_bytes = 0
            for start_idx, end_idx in ranges[:max_regions]:
                start_off = start_idx * chunk_size
                end_off = min(cur_size, (end_idx + 1) * chunk_size)
                length = max(0, end_off - start_off)

                old_list = (
                    base_hashes[start_idx : end_idx + 1] if start_idx < len(base_hashes) else []
                )
                new_list = (
                    cur_hashes[start_idx : end_idx + 1] if start_idx < len(cur_hashes) else []
                )

                raw = _read_region_preview(path, start_off, length, cap=preview_cap)
                hex_preview = raw.hex().upper()
                ascii_preview = "".join(chr(b) if 32 <= b < 127 else "." for b in raw)

                regions_out.append(
                    {
                        "chunk_idx_start": start_idx,
                        "chunk_idx_end": end_idx,
                        "start_offset": start_off,
                        "end_offset": end_off,
                        "length": length,
                        "old_chunk_hashes": old_list,
                        "new_chunk_hashes": new_list,
                        "preview": {"bytes": len(raw), "hex": hex_preview, "ascii": ascii_preview},
                    }
                )
                chunk_floor_bytes += length

            # ----- Estimate % changed -----
            # default: non-ZIP -> chunk-floor estimate
            approx_bytes_changed = chunk_floor_bytes
            estimation_method = "chunks"

            if zip_changes:
                # for Office containers, use only ZIP delta for the %.
                # this avoids "100%" when the whole archive gets re-packed.
                changed_zip_bytes = 0
                for z in zip_changes:
                    ch = (z.get("change") or "").lower()
                    if ch == "added":
                        changed_zip_bytes += int(z.get("size", 0))
                    elif ch == "removed":
                        changed_zip_bytes += int(z.get("size", 0))
                    elif ch == "modified":
                        sb = int(z.get("size_before", 0))
                        sa = int(z.get("size_after", 0))
                        delta = abs(sa - sb)
                        if delta == 0 and z.get("crc_before") != z.get("crc_after"):
                            # same size but different bytes -> treat as a tiny content tweak.
                            delta = min(sa, 512)  # cap the tiny edit at 1 KiB
                        changed_zip_bytes += delta

                # in case everything is metadata-only and delta rounds to zero, give a small floor.
                if changed_zip_bytes == 0 and changed_idxs:
                    changed_zip_bytes = min(cur_size, 1024)

                approx_bytes_changed = min(int(changed_zip_bytes), cur_size)
                estimation_method = "zip-members-delta"

            # initial byte-based percentage (will be refined if text_diff is available)
            denom = float(max(1, cur_size))
            percent = round((approx_bytes_changed / denom) * 100.0, 2)

            # ----- try to extract text-based diff for readable files -----
            text_diff: dict[str, Any] | None = None
            file_type = "binary"

            # check if this is a text file, Office document, or PDF
            is_text = _is_text_file(path)
            is_office = (path or "").lower().endswith((".docx", ".xlsx", ".pptx"))
            is_pdf = _is_pdf_file(path)

            if is_text or is_office or is_pdf:
                # check if we have a baseline blob stored
                baseline_blob = target.get("baseline_blob")
                baseline_text = None
                current_text = None

                if (
                    baseline_blob
                    and isinstance(baseline_blob, dict)
                    and not baseline_blob.get("error")
                ):
                    # read baseline from stored blob
                    stored_path = baseline_blob.get("stored_path")
                    if stored_path and os.path.exists(stored_path):
                        if is_office:
                            # extract text from Office document baseline
                            baseline_text, _ = _extract_text_from_office_doc(stored_path)
                        elif is_pdf:
                            # extract text from PDF baseline
                            baseline_text, _ = _extract_text_from_pdf(stored_path)
                        else:
                            # read as regular text file
                            baseline_text, _ = _read_text_file(stored_path)

                # read current file
                if is_office:
                    # extract text from Office document
                    current_text, text_error = _extract_text_from_office_doc(path)
                elif is_pdf:
                    # extract text from PDF
                    current_text, text_error = _extract_text_from_pdf(path)
                else:
                    # read as regular text file
                    current_text, text_error = _read_text_file(path)

                if baseline_text is not None and current_text is not None:
                    # both files readable, compute line diff
                    old_lines = baseline_text.splitlines(keepends=False)
                    new_lines = current_text.splitlines(keepends=False)
                    diff_segments = _compute_line_diff(old_lines, new_lines)

                    # extract images and formatting for Office documents
                    images_baseline: list[dict[str, Any]] = []
                    images_current: list[dict[str, Any]] = []
                    image_changes: list[dict[str, Any]] = []
                    formatting_changes: list[dict[str, Any]] = []

                    if is_office:
                        # extract images from both versions
                        images_baseline = (
                            _extract_images_from_office_doc(stored_path)
                            if stored_path and os.path.exists(stored_path)
                            else []
                        )
                        images_current = _extract_images_from_office_doc(path)

                        # compare images to detect add/remove/replace/resize
                        # create maps by hash for quick lookup
                        baseline_img_map = {img["hash"]: img for img in images_baseline}
                        current_img_map = {img["hash"]: img for img in images_current}

                        # also create maps by name for resize detection (when hash might be same but dimensions differ)
                        baseline_img_by_name = {img["name"]: img for img in images_baseline}
                        current_img_by_name = {img["name"]: img for img in images_current}

                        baseline_hashes = set(baseline_img_map.keys())
                        current_hashes = set(current_img_map.keys())

                        # added images (new hash)
                        for img_hash in current_hashes - baseline_hashes:
                            image_changes.append(
                                {
                                    "change": "added",
                                    "image": current_img_map[img_hash],
                                    "type": "image",
                                }
                            )

                        # removed images (hash no longer exists)
                        for img_hash in baseline_hashes - current_hashes:
                            image_changes.append(
                                {
                                    "change": "removed",
                                    "image": baseline_img_map[img_hash],
                                    "type": "image",
                                }
                            )

                        # check for resized images (same name/hash but different dimensions)
                        # use tolerance threshold to avoid false positives from minor dimension differences
                        # tolerance: >1 px or >1% change (prevents false positives from rounding/rendering differences)
                        RESIZE_TOLERANCE_PX = 1  # pixel tolerance
                        RESIZE_TOLERANCE_PCT = 0.01  # 1% tolerance

                        common_names = set(baseline_img_by_name.keys()) & set(
                            current_img_by_name.keys()
                        )
                        for img_name in common_names:
                            baseline_img = baseline_img_by_name[img_name]
                            current_img = current_img_by_name[img_name]

                            baseline_width = baseline_img.get("width")
                            baseline_height = baseline_img.get("height")
                            current_width = current_img.get("width")
                            current_height = current_img.get("height")

                            # check if dimensions changed beyond tolerance
                            if (
                                baseline_width is not None
                                and baseline_height is not None
                                and current_width is not None
                                and current_height is not None
                            ):
                                # calculate absolute and percentage differences
                                width_diff = abs(current_width - baseline_width)
                                height_diff = abs(current_height - baseline_height)
                                width_pct_diff = (
                                    width_diff / baseline_width if baseline_width > 0 else 0
                                )
                                height_pct_diff = (
                                    height_diff / baseline_height if baseline_height > 0 else 0
                                )

                                # check if change exceeds tolerance (>1 px or >1%)
                                is_resized = (
                                    width_diff > RESIZE_TOLERANCE_PX
                                    or width_pct_diff > RESIZE_TOLERANCE_PCT
                                ) or (
                                    height_diff > RESIZE_TOLERANCE_PX
                                    or height_pct_diff > RESIZE_TOLERANCE_PCT
                                )

                                if is_resized:
                                    # image was resized beyond tolerance threshold
                                    image_changes.append(
                                        {
                                            "change": "resized",
                                            "image": current_img,
                                            "old_width": baseline_width,
                                            "old_height": baseline_height,
                                            "new_width": current_width,
                                            "new_height": current_height,
                                            "type": "image",
                                        }
                                    )
                            # also check if hash is same but dimensions are missing in one version
                            elif baseline_img.get("hash") == current_img.get("hash"):
                                # same hash means same image file, but dimensions might have changed
                                # this is a resize case where we detected the same file but dimensions differ
                                if (baseline_width is not None or baseline_height is not None) and (
                                    current_width is not None or current_height is not None
                                ):
                                    # calculate differences if both dimensions are available
                                    if (
                                        baseline_width is not None
                                        and baseline_height is not None
                                        and current_width is not None
                                        and current_height is not None
                                    ):
                                        width_diff = abs(current_width - baseline_width)
                                        height_diff = abs(current_height - baseline_height)
                                        width_pct_diff = (
                                            width_diff / baseline_width if baseline_width > 0 else 0
                                        )
                                        height_pct_diff = (
                                            height_diff / baseline_height
                                            if baseline_height > 0
                                            else 0
                                        )

                                        # check if change exceeds tolerance
                                        is_resized = (
                                            width_diff > RESIZE_TOLERANCE_PX
                                            or width_pct_diff > RESIZE_TOLERANCE_PCT
                                        ) or (
                                            height_diff > RESIZE_TOLERANCE_PX
                                            or height_pct_diff > RESIZE_TOLERANCE_PCT
                                        )

                                        if is_resized:
                                            image_changes.append(
                                                {
                                                    "change": "resized",
                                                    "image": current_img,
                                                    "old_width": baseline_width,
                                                    "old_height": baseline_height,
                                                    "new_width": current_width,
                                                    "new_height": current_height,
                                                    "type": "image",
                                                }
                                            )
                                    else:
                                        # one dimension missing - treat as resize if dimensions differ
                                        if (
                                            baseline_width != current_width
                                            or baseline_height != current_height
                                        ):
                                            image_changes.append(
                                                {
                                                    "change": "resized",
                                                    "image": current_img,
                                                    "old_width": baseline_width,
                                                    "old_height": baseline_height,
                                                    "new_width": current_width,
                                                    "new_height": current_height,
                                                    "type": "image",
                                                }
                                            )

                        # add image changes to formatting_changes list so they appear in the formatting panel
                        # this ensures image changes are displayed alongside formatting changes
                        # keep the existing panel behavior (show 5, then "Show more") - don't change pagination
                        for img_change in image_changes:
                            change_type = img_change.get("change", "")
                            if change_type == "added":
                                formatting_changes.append(
                                    {
                                        "change": "added",
                                        "text": f"Image: {img_change.get('image', {}).get('name', 'Unknown')}",
                                        "type": "image",
                                        "image": img_change.get("image", {}),
                                    }
                                )
                            elif change_type == "removed":
                                formatting_changes.append(
                                    {
                                        "change": "removed",
                                        "text": f"Image: {img_change.get('image', {}).get('name', 'Unknown')}",
                                        "type": "image",
                                        "image": img_change.get("image", {}),
                                    }
                                )
                            elif change_type == "resized":
                                old_w = img_change.get("old_width")
                                old_h = img_change.get("old_height")
                                new_w = img_change.get("new_width")
                                new_h = img_change.get("new_height")
                                formatting_changes.append(
                                    {
                                        "change": "modified",
                                        "text": f"Image: {img_change.get('image', {}).get('name', 'Unknown')}",
                                        "type": "image",
                                        "image": img_change.get("image", {}),
                                        "changed_attrs": [
                                            f"size: {old_w}x{old_h} → {new_w}x{new_h}"
                                        ],
                                        "removed_attrs": [],
                                        "added_attrs": [],
                                    }
                                )

                        # extract and compare formatting, cross-reference with text diff to avoid false positives
                        # NEW IMPLEMENTATION: Proper tokenization, exact spans, no fragmentation, no unrelated entries
                        formatting_baseline_map = (
                            _extract_formatting_from_office_doc(stored_path)
                            if stored_path and os.path.exists(stored_path)
                            else {}
                        )
                        formatting_current_map = _extract_formatting_from_office_doc(path)

                        # helper function to normalize text for comparison (preserve word boundaries)
                        def normalize_text_for_matching(t: str) -> str:
                            """normalize text: lowercase, preserve word boundaries, remove extra whitespace"""
                            if not t:
                                return ""
                            # normalize whitespace but preserve word boundaries
                            normalized = " ".join(t.split()).lower().strip()
                            return normalized

                        # build precise map of unchanged text segments from diff
                        # ONLY collect from character-level diffs in modified segments
                        # This ensures we only detect formatting changes for text that actually appears unchanged
                        unchanged_text_segments = (
                            {}
                        )  # normalized -> set of original text fragments (exact text only)
                        removed_text_segments = set()  # normalized text that was removed
                        added_text_segments = set()  # normalized text that was added

                        # helper to add text fragment to unchanged segments
                        # Use proper tokenization to avoid fragmentation
                        def add_unchanged_text_fragment(text_fragment: str):
                            """add exact text fragment to unchanged segments, using proper tokenization"""
                            if not text_fragment or not isinstance(text_fragment, str):
                                return
                            text_fragment = text_fragment.strip()
                            if not text_fragment:
                                return

                            try:
                                # Tokenize the fragment to get complete words
                                tokens = _tokenize_text(text_fragment)

                                # Add each complete word/token
                                for token, start, end in tokens:
                                    if (
                                        token and len(token.strip()) >= 2
                                    ):  # Only add meaningful tokens (>= 2 chars)
                                        normalized = normalize_text_for_matching(token)
                                        if normalized and len(normalized) > 0:
                                            if normalized not in unchanged_text_segments:
                                                unchanged_text_segments[normalized] = set()
                                            # store the original token as-is (preserve case, spacing)
                                            unchanged_text_segments[normalized].add(token.strip())

                                # Also add the full fragment if it's substantial (>= 3 chars) to handle exact spans
                                if len(text_fragment) >= 3:
                                    normalized = normalize_text_for_matching(text_fragment)
                                    if normalized and len(normalized) > 0:
                                        if normalized not in unchanged_text_segments:
                                            unchanged_text_segments[normalized] = set()
                                        unchanged_text_segments[normalized].add(text_fragment)
                            except Exception:
                                # If tokenization fails, fall back to simple approach
                                if len(text_fragment) >= 2:
                                    normalized = normalize_text_for_matching(text_fragment)
                                    if normalized and len(normalized) > 0:
                                        if normalized not in unchanged_text_segments:
                                            unchanged_text_segments[normalized] = set()
                                        unchanged_text_segments[normalized].add(text_fragment)

                        # collect unchanged text from character-level diffs in modified segments
                        # also handle "equal" segments - they might have formatting-only changes
                        for seg in diff_segments:
                            if seg.get("type") == "modified":
                                # collect unchanged text from character-level diffs
                                old_char_diffs = seg.get("old_char_diffs", [])
                                new_char_diffs = seg.get("new_char_diffs", [])

                                if old_char_diffs or new_char_diffs:
                                    for i in range(max(len(old_char_diffs), len(new_char_diffs))):
                                        # collect equal parts from character-level diffs
                                        # these are text fragments that didn't change
                                        if i < len(old_char_diffs) and old_char_diffs[i]:
                                            for part in old_char_diffs[i]:
                                                if part.get("type") == "equal" and part.get("text"):
                                                    text = part.get("text", "").strip()
                                                    if text and len(text) > 0:
                                                        # only add if it's a meaningful fragment (at least 2 chars or a complete word)
                                                        # this prevents matching single characters like "e" incorrectly
                                                        if len(text) >= 2 or text.isalnum():
                                                            add_unchanged_text_fragment(text)

                                                # collect removed text to exclude
                                                # collect both the full text and individual words/tokens
                                                # this ensures we exclude formatting on any part of removed text
                                                if part.get("type") == "removed" and part.get(
                                                    "text"
                                                ):
                                                    text = part.get("text", "").strip()
                                                    if text:
                                                        # add the full normalized text
                                                        normalized = normalize_text_for_matching(
                                                            text
                                                        )
                                                        if normalized:
                                                            removed_text_segments.add(normalized)
                                                        # also add individual words/tokens from removed text
                                                        # this ensures we don't match formatting on words that were removed
                                                        import re

                                                        words = re.findall(r"\b\w+\b", text)
                                                        for word in words:
                                                            if len(word) >= 2:
                                                                word_norm = (
                                                                    normalize_text_for_matching(
                                                                        word
                                                                    )
                                                                )
                                                                if word_norm:
                                                                    removed_text_segments.add(
                                                                        word_norm
                                                                    )

                                        if i < len(new_char_diffs) and new_char_diffs[i]:
                                            for part in new_char_diffs[i]:
                                                if part.get("type") == "equal" and part.get("text"):
                                                    text = part.get("text", "").strip()
                                                    if text and len(text) > 0:
                                                        # only add if it's a meaningful fragment
                                                        if len(text) >= 2 or text.isalnum():
                                                            add_unchanged_text_fragment(text)

                                                # collect added text to exclude
                                                # collect both the full text and individual words/tokens
                                                # this ensures we exclude formatting on any part of added text
                                                if part.get("type") == "added" and part.get("text"):
                                                    text = part.get("text", "").strip()
                                                    if text:
                                                        # add the full normalized text
                                                        normalized = normalize_text_for_matching(
                                                            text
                                                        )
                                                        if normalized:
                                                            added_text_segments.add(normalized)
                                                        # also add individual words/tokens from added text
                                                        # this ensures we don't match formatting on words that were added
                                                        import re

                                                        words = re.findall(r"\b\w+\b", text)
                                                        for word in words:
                                                            if len(word) >= 2:
                                                                word_norm = (
                                                                    normalize_text_for_matching(
                                                                        word
                                                                    )
                                                                )
                                                                if word_norm:
                                                                    added_text_segments.add(
                                                                        word_norm
                                                                    )
                                else:
                                    # no character-level diffs - might be formatting-only change
                                    # if lines are the same (normalized), treat as formatting change
                                    old_lines_seg = seg.get("old_lines", [])
                                    new_lines_seg = seg.get("new_lines", [])
                                    if (
                                        old_lines_seg
                                        and new_lines_seg
                                        and len(old_lines_seg) == len(new_lines_seg)
                                    ):
                                        for old_line, new_line in zip(old_lines_seg, new_lines_seg):
                                            # normalize and compare - if text is same, it's formatting change
                                            old_norm = normalize_text_for_matching(old_line)
                                            new_norm = normalize_text_for_matching(new_line)
                                            if old_norm == new_norm and old_norm:
                                                # text is the same, only formatting changed
                                                # extract words from the line for matching
                                                import re

                                                words = re.findall(r"\b\w+\b", new_line)
                                                for word in words:
                                                    if len(word) >= 2:
                                                        add_unchanged_text_fragment(word)

                            elif seg.get("type") == "equal":
                                # "equal" segments might have formatting-only changes
                                # extract words from these lines for formatting comparison
                                old_lines_seg = seg.get("old_lines", [])
                                new_lines_seg = seg.get("new_lines", [])
                                # use new lines (they should match old lines for equal segments)
                                for line in new_lines_seg if new_lines_seg else old_lines_seg:
                                    if line:
                                        # extract individual words for matching
                                        import re

                                        words = re.findall(r"\b\w+\b", line)
                                        for word in words:
                                            if len(word) >= 2:
                                                add_unchanged_text_fragment(word)

                            elif seg.get("type") == "removed":
                                # text was removed - mark for exclusion
                                # collect both full lines and individual words to ensure comprehensive exclusion
                                old_lines_seg = seg.get("old_lines", [])
                                for line in old_lines_seg:
                                    if line:
                                        # add the full normalized line
                                        normalized = normalize_text_for_matching(line)
                                        if normalized:
                                            removed_text_segments.add(normalized)
                                        # also add individual words/tokens from removed lines
                                        # this ensures we don't match formatting on words that were removed
                                        import re

                                        words = re.findall(r"\b\w+\b", line)
                                        for word in words:
                                            if len(word) >= 2:
                                                word_norm = normalize_text_for_matching(word)
                                                if word_norm:
                                                    removed_text_segments.add(word_norm)

                            elif seg.get("type") == "added":
                                # text was added - mark for exclusion
                                # collect both full lines and individual words to ensure comprehensive exclusion
                                new_lines_seg = seg.get("new_lines", [])
                                for line in new_lines_seg:
                                    if line:
                                        # add the full normalized line
                                        normalized = normalize_text_for_matching(line)
                                        if normalized:
                                            added_text_segments.add(normalized)
                                        # also add individual words/tokens from added lines
                                        # this ensures we don't match formatting on words that were added
                                        import re

                                        words = re.findall(r"\b\w+\b", line)
                                        for word in words:
                                            if len(word) >= 2:
                                                word_norm = normalize_text_for_matching(word)
                                                if word_norm:
                                                    added_text_segments.add(word_norm)

                        # match formatting entries to unchanged text segments
                        # STRICT matching: only match exact text or complete words
                        import re

                        # helper function to check if a formatting text fragment matches unchanged text
                        # STRICT: only exact matches or complete word matches, no fragmentation
                        def formatting_text_matches_unchanged(
                            fmt_text: str, unchanged_segments: dict
                        ) -> tuple[bool, str | None]:
                            """check if formatting text fragment matches any unchanged text segment
                            returns (matches, matched_normalized_text)
                            STRICT: only exact matches or complete word matches, no fragmentation"""
                            if not fmt_text or not fmt_text.strip():
                                return False, None

                            normalized_fmt = normalize_text_for_matching(fmt_text)
                            if not normalized_fmt or len(normalized_fmt) < 1:
                                return False, None

                            # Tokenize the formatting text to get complete words
                            try:
                                fmt_tokens = _tokenize_text(fmt_text)
                                fmt_token_texts = [
                                    token
                                    for token, _, _ in fmt_tokens
                                    if token and len(token.strip()) >= 2
                                ]
                            except Exception:
                                # If tokenization fails, use empty list
                                fmt_token_texts = []

                            # Priority 1: exact match (after normalization) - highest priority
                            if normalized_fmt in unchanged_segments:
                                return True, normalized_fmt

                            # Priority 1.5: exact match of any complete token from formatting text
                            for fmt_token in fmt_token_texts:
                                normalized_token = normalize_text_for_matching(fmt_token)
                                if normalized_token in unchanged_segments:
                                    return True, normalized_token

                            # Priority 2: check if formatting text is a complete word in any unchanged segment
                            # Only match if it appears as a complete word (word boundaries)
                            fmt_no_space = (
                                normalized_fmt.replace(" ", "").replace("\n", "").replace("\t", "")
                            )

                            if not fmt_no_space or len(fmt_no_space) < 1:
                                return False, None

                            # Check all unchanged segments for word matches
                            # This handles cases where Office splits text into runs
                            # STRICT: Only match complete words, not fragments
                            for unchanged_norm, unchanged_orig_set in unchanged_segments.items():
                                # Check if formatting text is exactly equal to unchanged text (case-insensitive)
                                if normalized_fmt == unchanged_norm:
                                    return True, unchanged_norm

                                # Check if any complete token from formatting text matches unchanged text
                                for fmt_token in fmt_token_texts:
                                    normalized_token = normalize_text_for_matching(fmt_token)
                                    if normalized_token == unchanged_norm:
                                        return True, unchanged_norm

                                # Check if formatting text appears as a complete word in any original text
                                for orig_text in unchanged_orig_set:
                                    # Exact match (case-insensitive)
                                    if fmt_text.strip().lower() == orig_text.strip().lower():
                                        return True, unchanged_norm

                                    # Word boundary match - check if formatting text is a complete word
                                    # Only match if it's a complete word, not a fragment
                                    word_pattern = r"\b" + re.escape(fmt_text.strip()) + r"\b"
                                    if re.search(word_pattern, orig_text, re.IGNORECASE):
                                        return True, unchanged_norm

                                    # Also check if any complete token from formatting text matches
                                    for fmt_token in fmt_token_texts:
                                        token_pattern = r"\b" + re.escape(fmt_token.strip()) + r"\b"
                                        if re.search(token_pattern, orig_text, re.IGNORECASE):
                                            normalized_token = normalize_text_for_matching(
                                                fmt_token
                                            )
                                            return True, (
                                                normalized_token
                                                if normalized_token in unchanged_segments
                                                else unchanged_norm
                                            )

                            # For very short fragments (1-2 chars like "te"), be more permissive
                            # Check if it's an exact substring match (might be part of a word that was split)
                            if len(fmt_no_space) <= 2:
                                # For very short fragments, only match if it appears as a complete word
                                # or if it's an exact match to unchanged text
                                for (
                                    unchanged_norm,
                                    unchanged_orig_set,
                                ) in unchanged_segments.items():
                                    # Check exact match first
                                    if normalized_fmt == unchanged_norm:
                                        return True, unchanged_norm

                                    # Check if it's a complete word in any original text
                                    for orig_text in unchanged_orig_set:
                                        word_pattern = r"\b" + re.escape(fmt_text.strip()) + r"\b"
                                        if re.search(word_pattern, orig_text, re.IGNORECASE):
                                            return True, unchanged_norm

                                    # For very short fragments, also check if it's an exact substring
                                    # but only if the unchanged text is also very short (likely a fragment)
                                    unchanged_no_space = (
                                        unchanged_norm.replace(" ", "")
                                        .replace("\n", "")
                                        .replace("\t", "")
                                    )
                                    if fmt_no_space == unchanged_no_space:
                                        return True, unchanged_norm

                                return False, None

                            # Priority 3: for longer fragments (>=3 chars), check if they're complete words or exact substrings
                            best_match = None
                            best_match_score = 0.0

                            for unchanged_norm, unchanged_orig_set in unchanged_segments.items():
                                unchanged_no_space = (
                                    unchanged_norm.replace(" ", "")
                                    .replace("\n", "")
                                    .replace("\t", "")
                                )

                                if not unchanged_no_space:
                                    continue

                                # Check if formatting text is a complete word in the unchanged text
                                # This handles cases where Office splits "endpoint" into "end" and "point"
                                for orig_text in unchanged_orig_set:
                                    # First check exact match (case-insensitive)
                                    if fmt_text.strip().lower() == orig_text.strip().lower():
                                        # Exact match - always valid
                                        return True, unchanged_norm

                                    # Check if formatting text appears as a complete word
                                    word_pattern = r"\b" + re.escape(fmt_text.strip()) + r"\b"
                                    if re.search(word_pattern, orig_text, re.IGNORECASE):
                                        # It's a complete word match - high confidence
                                        return True, unchanged_norm

                                # Check if formatting text is contained in unchanged text as exact substring
                                # But only if it's a substantial portion (at least 50% for fragments >= 3 chars)
                                if fmt_no_space in unchanged_no_space and len(fmt_no_space) >= 3:
                                    overlap_ratio = (
                                        len(fmt_no_space) / len(unchanged_no_space)
                                        if unchanged_no_space
                                        else 0
                                    )
                                    # Only match if it's at least 50% of the unchanged text
                                    # This prevents matching "an" to "advanced" incorrectly
                                    if overlap_ratio >= 0.5:
                                        if overlap_ratio > best_match_score:
                                            best_match = unchanged_norm
                                            best_match_score = overlap_ratio

                                # Check if unchanged text is contained in formatting text (formatting is larger)
                                # This handles cases where Office combines multiple runs
                                elif (
                                    unchanged_no_space in fmt_no_space
                                    and len(unchanged_no_space) >= 3
                                ):
                                    overlap_ratio = (
                                        len(unchanged_no_space) / len(fmt_no_space)
                                        if fmt_no_space
                                        else 0
                                    )
                                    # Only match if unchanged text is at least 70% of formatting text
                                    if overlap_ratio >= 0.7:
                                        if overlap_ratio > best_match_score:
                                            best_match = unchanged_norm
                                            best_match_score = overlap_ratio

                            # Return best match only if it meets strict criteria
                            if best_match and best_match_score >= 0.5:
                                return True, best_match

                            return False, None

                        # build maps: normalized unchanged text -> list of formatting entries that match it
                        unchanged_to_baseline_fmt = (
                            {}
                        )  # normalized_text -> list of (orig_text, fmt_info)
                        unchanged_to_current_fmt = (
                            {}
                        )  # normalized_text -> list of (orig_text, fmt_info)

                        # also track which formatting entries we've matched (to avoid duplicates)
                        matched_baseline_entries = set()
                        matched_current_entries = set()

                        # helper function to check if formatting text contains any removed/added words
                        def formatting_contains_removed_or_added(fmt_text: str) -> bool:
                            """check if formatting text contains any words that were removed or added
                            returns True if any word in formatting text is in removed/added segments
                            """
                            if not fmt_text or not fmt_text.strip():
                                return False

                            # normalize the full formatting text
                            normalized_fmt = normalize_text_for_matching(fmt_text)
                            if (
                                normalized_fmt in removed_text_segments
                                or normalized_fmt in added_text_segments
                            ):
                                return True

                            # check individual words in formatting text
                            import re

                            words = re.findall(r"\b\w+\b", fmt_text)
                            for word in words:
                                if len(word) >= 2:
                                    word_norm = normalize_text_for_matching(word)
                                    if word_norm and (
                                        word_norm in removed_text_segments
                                        or word_norm in added_text_segments
                                    ):
                                        return True

                            return False

                        # match baseline formatting to unchanged text
                        for fmt_text, fmt_info in formatting_baseline_map.items():
                            if not fmt_text or not fmt_text.strip():
                                continue

                            # skip if we've already matched this entry
                            if fmt_text in matched_baseline_entries:
                                continue

                            # CRITICAL: skip if formatting text contains any removed/added words
                            # this ensures content changes don't trigger formatting changes
                            if formatting_contains_removed_or_added(fmt_text):
                                continue

                            matches, matched_normalized = formatting_text_matches_unchanged(
                                fmt_text, unchanged_text_segments
                            )
                            if matches and matched_normalized:
                                # skip if this text was removed or added
                                if (
                                    matched_normalized in removed_text_segments
                                    or matched_normalized in added_text_segments
                                ):
                                    continue

                                if matched_normalized not in unchanged_to_baseline_fmt:
                                    unchanged_to_baseline_fmt[matched_normalized] = []
                                unchanged_to_baseline_fmt[matched_normalized].append(
                                    (fmt_text, fmt_info)
                                )
                                matched_baseline_entries.add(fmt_text)

                        # match current formatting to unchanged text
                        for fmt_text, fmt_info in formatting_current_map.items():
                            if not fmt_text or not fmt_text.strip():
                                continue

                            # skip if we've already matched this entry
                            if fmt_text in matched_current_entries:
                                continue

                            # CRITICAL: skip if formatting text contains any removed/added words
                            # this ensures content changes don't trigger formatting changes
                            if formatting_contains_removed_or_added(fmt_text):
                                continue

                            matches, matched_normalized = formatting_text_matches_unchanged(
                                fmt_text, unchanged_text_segments
                            )
                            if matches and matched_normalized:
                                # skip if this text was removed or added
                                if (
                                    matched_normalized in removed_text_segments
                                    or matched_normalized in added_text_segments
                                ):
                                    continue

                                if matched_normalized not in unchanged_to_current_fmt:
                                    unchanged_to_current_fmt[matched_normalized] = []
                                unchanged_to_current_fmt[matched_normalized].append(
                                    (fmt_text, fmt_info)
                                )
                                matched_current_entries.add(fmt_text)

                        # fallback: if formatting exists in both baseline and current for the same text,
                        # but we didn't match it to unchanged segments, try to match it anyway
                        # (this handles cases where the diff algorithm didn't capture the text correctly)
                        # CRITICAL: only match if text doesn't contain removed/added words
                        common_formatting_texts = set(formatting_baseline_map.keys()) & set(
                            formatting_current_map.keys()
                        )
                        for fmt_text in common_formatting_texts:
                            # skip if already matched
                            if (
                                fmt_text in matched_baseline_entries
                                and fmt_text in matched_current_entries
                            ):
                                continue

                            # CRITICAL: skip if formatting text contains any removed/added words
                            # this ensures content changes don't trigger formatting changes
                            if formatting_contains_removed_or_added(fmt_text):
                                continue

                            normalized_fmt = normalize_text_for_matching(fmt_text)
                            if not normalized_fmt:
                                continue

                            # skip if this text was removed or added
                            if (
                                normalized_fmt in removed_text_segments
                                or normalized_fmt in added_text_segments
                            ):
                                continue

                            # check if this text appears in any unchanged segment (try to match it)
                            matches, matched_normalized = formatting_text_matches_unchanged(
                                fmt_text, unchanged_text_segments
                            )
                            if matches and matched_normalized:
                                # match found - add to maps
                                if matched_normalized not in unchanged_to_baseline_fmt:
                                    unchanged_to_baseline_fmt[matched_normalized] = []
                                if fmt_text not in matched_baseline_entries:
                                    unchanged_to_baseline_fmt[matched_normalized].append(
                                        (fmt_text, formatting_baseline_map[fmt_text])
                                    )
                                    matched_baseline_entries.add(fmt_text)

                                if matched_normalized not in unchanged_to_current_fmt:
                                    unchanged_to_current_fmt[matched_normalized] = []
                                if fmt_text not in matched_current_entries:
                                    unchanged_to_current_fmt[matched_normalized].append(
                                        (fmt_text, formatting_current_map[fmt_text])
                                    )
                                    matched_current_entries.add(fmt_text)

                        # process each unchanged text segment that has formatting
                        # only process segments where we have formatting entries that match
                        processed_segments = set()

                        # get all segments that have formatting matches
                        segments_with_formatting = set(unchanged_to_baseline_fmt.keys()) | set(
                            unchanged_to_current_fmt.keys()
                        )

                        for normalized_text in segments_with_formatting:
                            # skip if this text was removed or added
                            if (
                                normalized_text in removed_text_segments
                                or normalized_text in added_text_segments
                            ):
                                continue

                            # skip if this text is not in unchanged_text_segments (shouldn't happen, but safety check)
                            if normalized_text not in unchanged_text_segments:
                                continue

                            # skip if we've already processed this segment
                            if normalized_text in processed_segments:
                                continue

                            original_texts = unchanged_text_segments[normalized_text]

                            # get formatting for this segment
                            baseline_entries = unchanged_to_baseline_fmt.get(normalized_text, [])
                            current_entries = unchanged_to_current_fmt.get(normalized_text, [])

                            # only process if we have formatting in at least one version
                            if baseline_entries or current_entries:
                                processed_segments.add(normalized_text)

                                # merge formatting from all matching entries
                                baseline_merged_styles = {}
                                current_merged_styles = {}

                                # merge baseline formatting from all matching entries
                                # collect ALL attributes from ALL occurrences to ensure we don't miss any
                                # this aggregates contiguous tokens with the same formatting, regardless of baseline inheritance
                                # e.g., if "and" has purple on "an" and no color on "d", we collect both
                                # when matching to unchanged text "and", we aggregate formatting from both "an" and "d" entries
                                for orig_text, fmt_info in baseline_entries:
                                    if not fmt_info:
                                        continue
                                    styles = fmt_info.get("styles", {})
                                    # add all attributes from this entry
                                    for key, val in styles.items():
                                        if key not in baseline_merged_styles:
                                            baseline_merged_styles[key] = set()
                                        # add the value - this ensures we capture color, strikethrough, italic, etc.
                                        # this aggregates formatting across contiguous tokens that share the same baseline formatting
                                        baseline_merged_styles[key].add(val)
                                    # also check if there are attributes that should be present but aren't
                                    # (this handles cases where formatting was removed)

                                # merge current formatting from all matching entries
                                # collect ALL attributes from ALL occurrences to ensure we don't miss any
                                # this aggregates contiguous tokens with the same formatting, regardless of baseline inheritance
                                # e.g., if "and" has strikethrough on the whole word, we aggregate it across all matching entries
                                # even if baseline had purple only on "an" - the strikethrough should cover the entire word
                                for orig_text, fmt_info in current_entries:
                                    if not fmt_info:
                                        continue
                                    styles = fmt_info.get("styles", {})
                                    # add all attributes from this entry
                                    for key, val in styles.items():
                                        if key not in current_merged_styles:
                                            current_merged_styles[key] = set()
                                        # add the value - this ensures we capture color, strikethrough, italic, etc.
                                        # this aggregates formatting across contiguous tokens that share the same new formatting
                                        # regardless of how the baseline formatting was split (e.g., purple on "an" only)
                                        # the key fix: when strikethrough is applied to the whole word "and",
                                        # we aggregate it across all matching entries, so it covers the entire word
                                        # not just the "an" portion that had purple in baseline
                                        current_merged_styles[key].add(val)

                                # compare formatting and detect changes
                                # normalize style values before comparison to prevent phantom changes
                                # only compare on aligned token spans (unchanged text) to avoid false positives
                                changed_attrs = []
                                removed_attrs = []
                                added_attrs = []

                                # get all attribute keys from both versions
                                all_attr_keys = set(baseline_merged_styles.keys()) | set(
                                    current_merged_styles.keys()
                                )

                                for attr_key in all_attr_keys:
                                    baseline_vals = baseline_merged_styles.get(attr_key, set())
                                    current_vals = current_merged_styles.get(attr_key, set())

                                    # normalize values for comparison - prevents phantom changes from value format differences
                                    # e.g., "#A02B93" vs "A02B93" should be treated as the same color
                                    baseline_normalized = {
                                        _normalize_style_value(attr_key, val)
                                        for val in baseline_vals
                                    }
                                    current_normalized = {
                                        _normalize_style_value(attr_key, val)
                                        for val in current_vals
                                    }

                                    # remove empty strings from normalized sets (they represent "not set")
                                    baseline_normalized = {v for v in baseline_normalized if v}
                                    current_normalized = {v for v in current_normalized if v}

                                    if baseline_normalized and current_normalized:
                                        # attribute exists in both versions
                                        if baseline_normalized == current_normalized:
                                            # same normalized values - no change (prevents phantom changes)
                                            continue
                                        else:
                                            # different normalized values - changed
                                            # use original values for display (before normalization)
                                            old_val = next(iter(baseline_vals))
                                            new_val = next(iter(current_vals))
                                            # for color, show color names only (hex as fallback)
                                            if attr_key == "color":
                                                old_color_name = _hex_to_color_name(old_val)
                                                new_color_name = _hex_to_color_name(new_val)
                                                old_hex = (
                                                    f"#{old_val}"
                                                    if not old_val.startswith("#")
                                                    else old_val
                                                )
                                                new_hex = (
                                                    f"#{new_val}"
                                                    if not new_val.startswith("#")
                                                    else new_val
                                                )
                                                # Format as "color: purple → red" or use hex if no name found
                                                old_display = (
                                                    old_color_name if old_color_name else old_hex
                                                )
                                                new_display = (
                                                    new_color_name if new_color_name else new_hex
                                                )
                                                changed_attrs.append(
                                                    f"color: {old_display} → {new_display}"
                                                )
                                            else:
                                                changed_attrs.append(
                                                    f"{attr_key}: {old_val} → {new_val}"
                                                )
                                    elif baseline_normalized and not current_normalized:
                                        # attribute removed (e.g., bold was removed)
                                        # use original value for display
                                        old_val = next(iter(baseline_vals))
                                        # For color, show the removed color name (hex as fallback)
                                        if attr_key == "color":
                                            color_name = _hex_to_color_name(old_val)
                                            hex_display = (
                                                f"#{old_val}"
                                                if not old_val.startswith("#")
                                                else old_val
                                            )
                                            color_display = (
                                                color_name if color_name else hex_display
                                            )
                                            removed_attrs.append(f"color {color_display}")
                                        else:
                                            # For boolean attributes, just show the name
                                            removed_attrs.append(attr_key)
                                    elif not baseline_normalized and current_normalized:
                                        # attribute added (e.g., bold was added, or color was added)
                                        # use original value for display
                                        new_val = next(iter(current_vals))
                                        if attr_key == "color":
                                            # Format as "color name" - show color name only (hex as fallback)
                                            color_name = _hex_to_color_name(new_val)
                                            hex_display = (
                                                f"#{new_val}"
                                                if not new_val.startswith("#")
                                                else new_val
                                            )
                                            color_display = (
                                                color_name if color_name else hex_display
                                            )
                                            added_attrs.append(f"color {color_display}")
                                        else:
                                            # For boolean attributes, just show the name (no =true)
                                            added_attrs.append(attr_key)

                                # only report if there are actual formatting changes
                                if changed_attrs or removed_attrs or added_attrs:
                                    # choose the best display text - prefer the longest, most complete fragment
                                    # this ensures we show "endpoint" instead of "end" or "point"
                                    display_text = None
                                    max_len = 0
                                    for orig_text in original_texts:
                                        text_len = len(orig_text.strip())
                                        if text_len > max_len:
                                            max_len = text_len
                                            display_text = orig_text.strip()

                                    # fallback to normalized text if no original found
                                    if not display_text:
                                        display_text = normalized_text

                                    # CRITICAL: Filter out invalid/confused formatting diffs
                                    # If the detected formatting change does not map to clean, real token boundaries
                                    # (i.e., it looks merged, concatenated, or nonsense), discard it entirely
                                    import re

                                    # Check for common patterns that indicate invalid/confused formatting:
                                    # 1. Very long concatenated text without spaces (likely merged tokens)
                                    # 2. Text that contains multiple words concatenated without spaces
                                    # 3. Text that doesn't look like clean token boundaries

                                    # Remove all whitespace and check if it's too long without word boundaries
                                    text_no_space = re.sub(r"\s+", "", display_text)

                                    # Check if text looks like merged/concatenated nonsense
                                    # Pattern examples: "transparent."-readableaddtheirpytestFastAPIapp"
                                    # This has punctuation followed by concatenated words without spaces
                                    is_invalid = False

                                    # Check for patterns where words are concatenated without spaces
                                    # Look for sequences of lowercase followed by uppercase (camelCase-like but without spaces)
                                    # This catches patterns like "readableaddtheirpytestFastAPIapp"
                                    if len(display_text) > 10:
                                        # Check for camelCase-like patterns that suggest concatenation
                                        # Pattern: lowercase letters followed by uppercase (word boundaries without spaces)
                                        camel_case_matches = re.findall(
                                            r"[a-z]+[A-Z][a-z]+", display_text
                                        )
                                        if len(camel_case_matches) >= 2:
                                            # Multiple camelCase patterns - likely concatenated words
                                            is_invalid = True

                                    # Also check for very long runs of letters without spaces (likely concatenated)
                                    # Pattern: text that has > 20 consecutive letters without spaces
                                    if not is_invalid and len(text_no_space) > 20:
                                        # Count word-like sequences (sequences of letters)
                                        word_sequences = re.findall(r"[a-zA-Z]+", display_text)
                                        if len(word_sequences) >= 3:
                                            # Multiple word sequences - check if they're concatenated
                                            # If there are 3+ word sequences but no spaces between them,
                                            # it's likely concatenated nonsense
                                            # Check if the text has punctuation that suggests merging
                                            has_punctuation = bool(
                                                re.search(r'[.,;:!?\-"\'()]', display_text)
                                            )
                                            if has_punctuation and len(word_sequences) >= 3:
                                                # Has punctuation and multiple word sequences - likely merged text
                                                # Check if words are directly concatenated (no spaces between word sequences)
                                                # by checking if text length without spaces is close to sum of word lengths
                                                total_word_length = sum(
                                                    len(w) for w in word_sequences
                                                )
                                                # If text without spaces is close to word lengths, they're concatenated
                                                if abs(len(text_no_space) - total_word_length) < 5:
                                                    # Words are concatenated without spaces - invalid
                                                    is_invalid = True

                                    # Skip invalid formatting diffs
                                    if is_invalid:
                                        continue

                                    # build style lists for display (removed - no longer needed)
                                    # We now use changed_attrs, removed_attrs, added_attrs only
                                    # No need for old_styles/new_styles with =true/false noise
                                    old_style_list = []
                                    new_style_list = []

                                    formatting_changes.append(
                                        {
                                            "change": "modified",
                                            "text": display_text,
                                            "old_styles": old_style_list,
                                            "new_styles": new_style_list,
                                            "changed_attrs": changed_attrs,
                                            "removed_attrs": removed_attrs,
                                            "added_attrs": added_attrs,
                                            "type": "formatting",
                                        }
                                    )

                    # show entire file, no truncation
                    # the UI will handle scrolling for large files

                    text_diff = {
                        "type": "text",
                        "file_type": _nl_file_kind(path),
                        "is_config": _is_config_file(path),
                        "is_office": is_office,
                        "is_pdf": is_pdf,
                        "old_line_count": len(old_lines),
                        "new_line_count": len(new_lines),
                        "diff_segments": diff_segments,
                        "show_full_file": True,  # flag to show entire file
                        "images_baseline": images_baseline,
                        "images_current": images_current,
                        "image_changes": image_changes,
                        "formatting_changes": formatting_changes,
                    }
                    file_type = "text"

                    # recalculate percentage based on actual changed lines (more accurate for text files)
                    total_lines = max(len(old_lines), len(new_lines))
                    changed_lines = 0
                    changed_chars = 0
                    total_chars = 0

                    for seg in diff_segments:
                        seg_type = seg.get("type", "")
                        if seg_type == "added":
                            changed_lines += len(seg.get("new_lines", []))
                            for line in seg.get("new_lines", []):
                                changed_chars += len(line)
                        elif seg_type == "removed":
                            changed_lines += len(seg.get("old_lines", []))
                            for line in seg.get("old_lines", []):
                                changed_chars += len(line)
                        elif seg_type == "modified":
                            # for modified lines, count actual character changes
                            old_char_diffs = seg.get("old_char_diffs", [])
                            new_char_diffs = seg.get("new_char_diffs", [])
                            old_lines_seg = seg.get("old_lines", [])
                            new_lines_seg = seg.get("new_lines", [])

                            if old_char_diffs or new_char_diffs:
                                # count actual changed characters from character-level diffs
                                for i in range(max(len(old_lines_seg), len(new_lines_seg))):
                                    if i < len(old_char_diffs) and old_char_diffs[i]:
                                        for part in old_char_diffs[i]:
                                            if part.get("type") == "removed":
                                                changed_chars += len(part.get("text", ""))
                                    if i < len(new_char_diffs) and new_char_diffs[i]:
                                        for part in new_char_diffs[i]:
                                            if part.get("type") == "added":
                                                changed_chars += len(part.get("text", ""))
                                    changed_lines += 0.5  # partial line (weighted)
                            else:
                                # fallback: count all modified lines
                                changed_lines += max(len(old_lines_seg), len(new_lines_seg))
                                # estimate character changes
                                for line in old_lines_seg + new_lines_seg:
                                    changed_chars += len(line) // 2  # estimate 50% changed

                    # count total characters for percentage calculation
                    for line in old_lines + new_lines:
                        total_chars += len(line)

                    if total_lines > 0 and total_chars > 0:
                        # use character-based percentage for more accuracy (weighted 60%)
                        char_percent = (changed_chars / total_chars) * 100.0
                        # use line-based percentage (weighted 40%)
                        line_percent = (changed_lines / total_lines) * 100.0
                        # blend both
                        percent = round((char_percent * 0.6) + (line_percent * 0.4), 2)
                    elif total_lines > 0:
                        # fallback to line-based only
                        line_percent = (changed_lines / total_lines) * 100.0
                        byte_percent = (approx_bytes_changed / float(max(1, cur_size))) * 100.0
                        percent = round((line_percent * 0.7) + (byte_percent * 0.3), 2)

                    # cap at 100% and ensure non-negative
                    percent = min(100.0, max(0.0, percent))
                elif current_text is not None:
                    # only current file readable, baseline might be missing or binary
                    text_diff = {
                        "type": "text",
                        "file_type": _nl_file_kind(path),
                        "is_config": _is_config_file(path),
                        "is_office": is_office,
                        "old_line_count": 0,
                        "new_line_count": len(current_text.splitlines(keepends=False)),
                        "diff_segments": [],
                        "note": "Baseline file not available for comparison",
                    }
                    file_type = "text"
            elif zip_changes:
                file_type = "office"

            # generate summary text with accurate percentage
            nl = _nl_summary_for_diff(
                path=path,
                approx_bytes_changed=approx_bytes_changed,
                percent=percent,
                ranges=ranges,
                zip_changes=zip_changes,
                chunk_size=chunk_size,
                base_size=base_size,
                cur_size=cur_size,
            )

            out = {
                "ok": True,
                "file_type": file_type,
                "summary": {
                    "file": path,
                    "changed_chunks": len(changed_idxs),
                    "regions_returned": len(regions_out),
                    "approx_changed_bytes": approx_bytes_changed,
                    "approx_percent_of_file": percent,
                    "chunk_size": chunk_size,
                    "baseline_size": base_size,
                    "current_size": cur_size,
                },
                "estimation_method": estimation_method,  # "zip-members-delta" or "chunks"
                "summary_text": {
                    "headline": nl.get("headline", ""),
                    "details": nl.get("details", ""),
                },
                "zip_changes": zip_changes,
                "regions": regions_out,
                "text_diff": text_diff,  # new: text-based diff for readable files
            }
            return jsonify(out)

        except Exception as e:
            return jsonify({"ok": False, "error": str(e)}), 500

    # integrity baseline accept API: accept current file state as new baseline
    @app.post("/api/integrity/baseline/accept")
    @require_auth
    def api_integrity_baseline_accept():
        """
        accept current file state as the new baseline.
        this updates the baseline hash/mtime+size and creates a new baseline blob.
        for deleted files, marks the deletion as intentional and removes from watch list.
        """
        try:
            body: dict[str, Any] = cast(dict[str, Any], request.get_json(force=True) or {})
            path = str(body.get("path") or "").strip()
            if not path:
                return jsonify({"ok": False, "error": "missing path"}), 400

            rows = _read_integrity_targets()
            target = next(
                (r for r in rows if str(r.get("path") or "").lower() == path.lower()), None
            )
            if target is None:
                return jsonify({"ok": False, "error": "not on watch list"}), 400

            # check if file is deleted
            if not os.path.exists(_norm_user_path(path)):
                # file is deleted, mark as intentionally deleted and remove from watch list
                rule = str(target.get("rule") or "sha256").lower()

                # update existing Live Events entries for this path to reflect acceptance
                # append acceptance text to existing CRITICAL entry instead of creating new event
                _update_live_events_for_path(path, "✔ Change accepted (deletion accepted)")

                # note: we do not emit a new event for OK statuses, we update the existing CRITICAL entry instead
                # the file is removed from the watch list, so no integrity target update is needed

                # remove from watch list
                rows = [r for r in rows if str(r.get("path") or "").lower() != path.lower()]
                _write_integrity_targets(rows)
                return jsonify(
                    {
                        "ok": True,
                        "message": "File deletion accepted - removed from watch list",
                        "file_deleted": True,
                    }
                )

            rule = str(target.get("rule") or "sha256").lower()

            # update baseline based on rule
            if rule == "sha256":
                info = _sha256_file(path)
                if info.get("sha256"):
                    target["sha256"] = info["sha256"]
                    target["last_result"] = "OK (baseline updated)"

                    # update baseline blob
                    if ALLOW_MTIME_SNAPSHOTS:
                        try:
                            target["baseline_blob"] = _snapshot_file_to_cas(path)
                            _maybe_prune_baselines()
                        except Exception as e:
                            target["baseline_blob"] = {"error": str(e)}
                else:
                    return jsonify({"ok": False, "error": info.get("error", "hash failed")}), 400
            else:
                # mtime+size rule
                try:
                    st = os.stat(_norm_user_path(path))
                    target["mtime"] = int(st.st_mtime)
                    target["size"] = int(st.st_size)
                    target["last_result"] = "OK (baseline updated)"

                    # update baseline blob
                    if ALLOW_MTIME_SNAPSHOTS:
                        try:
                            target["baseline_blob"] = _snapshot_file_to_cas(path)
                            _maybe_prune_baselines()
                        except Exception as e:
                            target["baseline_blob"] = {"error": str(e)}
                except Exception as e:
                    return jsonify({"ok": False, "error": str(e)}), 400

            # update chunk baseline
            info_chunks = _chunk_hashes(path)
            if "error" not in info_chunks:
                target["chunks"] = info_chunks

            # update ZIP manifest if applicable
            zman = _zip_manifest(path)
            if zman:
                target["zip_manifest"] = zman
            else:
                target.pop("zip_manifest", None)

            _write_integrity_targets(rows)

            # update existing Live Events entries for this path to reflect acceptance
            # append acceptance text to existing CRITICAL entry instead of creating new event
            _update_live_events_for_path(path, "✔ Marked Safe")

            # note: we do not emit a new event for OK statuses, we update the existing CRITICAL entry instead
            # the integrity target status is already updated above with target["last_result"] = "OK (baseline updated)"

            return jsonify(
                {
                    "ok": True,
                    "message": "Baseline updated successfully",
                    "last_result": target.get("last_result", "OK (baseline updated)"),
                }
            )

        except Exception as e:
            return jsonify({"ok": False, "error": str(e)}), 500

    # integrity baseline download API: download stored baseline snapshot file
    @app.get("/api/integrity/baseline/download")
    @require_auth
    def api_integrity_baseline_download():
        """
        download the stored baseline blob for a given path, if present and valid.
        query: ?path=<watch-list path>
        """
        try:
            path = (request.args.get("path") or "").strip()
            if not path:
                return jsonify({"ok": False, "error": "missing path"}), 400

            rows = _read_integrity_targets()
            target = next(
                (r for r in rows if str(r.get("path") or "").lower() == path.lower()), None
            )
            if not target:
                return jsonify({"ok": False, "error": "not on watch list"}), 404

            blob = target.get("baseline_blob")
            if not isinstance(blob, dict) or blob.get("error"):
                return jsonify({"ok": False, "error": "no baseline blob available"}), 404

            if not _validate_cas_blob(blob):
                return jsonify({"ok": False, "error": "baseline blob failed validation"}), 409

            p = blob.get("stored_path")
            fname = os.path.basename(p) if p else "baseline.bin"
            return send_file(p, as_attachment=True, download_name=fname)
        except Exception as e:
            return jsonify({"ok": False, "error": str(e)}), 500

    # integrity browse API: Windows-only file picker dialog for selecting files to watch
    @app.post("/api/integrity/browse")
    @require_auth
    def api_integrity_browse():
        """
        windows-only file picker via tkinter. returns {"path": "..."} or {} if cancelled.
        we keep it very small and guarded. it will no-op on non-Windows.
        sets app icon and enables DPI awareness for crisp rendering.
        """
        try:
            if sys.platform != "win32":
                return jsonify({"error": "browse supported on Windows only"}), 400

            # enable DPI awareness for crisp rendering (must be done before creating any windows)
            try:
                import ctypes

                # try to set Per-Monitor DPI awareness (Windows 10+)
                try:
                    # PROCESS_PER_MONITOR_DPI_AWARE = 2
                    ctypes.windll.shcore.SetProcessDpiAwareness(2)
                except (AttributeError, OSError):
                    # fallback to system DPI awareness (Windows Vista+)
                    try:
                        # PROCESS_DPI_AWARE = 1
                        ctypes.windll.user32.SetProcessDPIAware()
                    except (AttributeError, OSError):
                        pass  # older Windows or already set
            except Exception:
                pass  # non-critical, continue without DPI awareness

            # late imports to avoid importing Tk on non-Windows
            import tkinter as _tk  # type: ignore
            from tkinter import filedialog as _fd  # type: ignore

            root = _tk.Tk()
            root.withdraw()

            # set app icon for dialog title bar and taskbar
            icon_path = BASE_DIR / "assets" / "favicon.ico"
            if icon_path.exists():
                try:
                    root.iconbitmap(str(icon_path))
                except Exception:
                    pass  # non-critical if icon fails to load

            root.attributes("-topmost", True)  # bring dialog front
            root.title("CustosEye - Select File")  # set window title

            sel = _fd.askopenfilename(title="Select a file to watch")
            try:
                root.destroy()
            except Exception:
                pass
            if sel:
                return jsonify({"path": sel})
            return jsonify({})
        except Exception as e:
            return jsonify({"error": str(e)}), 500

    return app


# global shutdown flag for graceful shutdown
_shutdown_requested = False


# run the dashboard: start the Flask app with optional Waitress server
def run_dashboard(event_bus) -> None:
    global _shutdown_requested
    app = build_app(event_bus)
    if app is None:
        raise RuntimeError(
            "build_app() returned None; check for exceptions or missing `return app`."
        )
    if HAVE_WAITRESS:
        # waitress doesn't have a built-in shutdown, so we'll use os._exit
        # but first, set up a shutdown endpoint that can be called
        try:
            _serve(app, host=CFG.host, port=CFG.port)
        except SystemExit:
            pass  # expected when shutting down
        except KeyboardInterrupt:
            pass  # expected when shutting down
    else:
        try:
            app.run(host=CFG.host, port=CFG.port, debug=False)
        except SystemExit:
            pass  # expected when shutting down
        except KeyboardInterrupt:
            pass  # expected when shutting down


# standalone mode (optional): if you run "python -m dashboard.app"
# creates a minimal event bus and starts agents + dashboard for testing
if __name__ == "__main__":
    # minimal pub/sub bus for standalone dashboard
    import queue

    class FanoutEventBus:
        def __init__(self) -> None:
            self._subs: list[queue.Queue] = []
            self._lock = threading.Lock()

        def publish(self, event: dict[str, Any]) -> None:
            with self._lock:
                subs = list(self._subs)
            for q in subs:
                try:
                    q.put_nowait(event)
                except queue.Full:
                    pass

        def subscribe(self):
            q: queue.Queue = queue.Queue(maxsize=1000)
            with self._lock:
                self._subs.append(q)

            def _iter():
                while True:
                    try:
                        yield q.get(timeout=0.2)
                    except queue.Empty:
                        yield None

            return _iter()

    bus = FanoutEventBus()

    # start agents (standalone)
    from agent.integrity_check import IntegrityChecker
    from agent.monitor import ProcessMonitor
    from agent.network_scan import NetworkSnapshot

    for target in (
        ProcessMonitor(publish=bus.publish).run,
        NetworkSnapshot(publish=bus.publish).run,
        IntegrityChecker(
            targets_path=INTEGRITY_TARGETS_PATH,
            publish=bus.publish,
            interval_sec=5.0,
        ).run,
    ):
        threading.Thread(target=target, daemon=True).start()

    run_dashboard(bus)
